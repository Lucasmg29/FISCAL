import tkinter as tk
from tkinter import messagebox
from tkinter import scrolledtext
import pyperclip
import subprocess
import shlex
import time
import os
import win32com.client
from pywinauto.application import Application
import psutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime



def fazer_login():
    caminho_executavel_sap = r'C:\Program Files\SAP\FrontEnd\SAPGUI\saplogon.exe'
    if not os.path.exists(caminho_executavel_sap):
        caminho_executavel_sap = r'C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe'
    
    print(f"Caminho do executável: {caminho_executavel_sap}")
    
    if not os.path.exists(caminho_executavel_sap):
        print("O arquivo não foi encontrado. Verifique o caminho.")
        messagebox.showerror("Erro", "O executável do SAP GUI não foi encontrado. Verifique o caminho.")
        return

    # Use shlex.split para tratar corretamente o caminho com espaços
    comando = shlex.split(f'"{caminho_executavel_sap}"')
    print(f"Comando para execução: {comando}")

    try:
        subprocess.Popen(comando)
    except Exception as e:
        print(f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        return

    time.sleep(3)

    usuario = entry_usuario.get()
    senha = entry_senha.get()
    centro = entry_centro.get()

    try:
        subprocess.Popen(comando)
    except Exception as e:
        print(f"Ocorreu um erro ao iniciar o SAP GUI: {e}")

    time.sleep(3)
    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    try:
        connection = application.OpenConnection("NOVO PVR - Produção SAP ECC", True)
    except Exception as e:
        print(f"Conexão 'PVR - Produção (Externo)' não encontrada. Tentando 'PVR - Produção (Interno)'...")
        try:
            connection = application.OpenConnection("NOVO PVR - Produção SAP ECC", True)
        except Exception as e:
            print(f"Ocorreu um erro ao abrir a conexão: {e}")

    session = connection.Children(0)

    try:
        # Preencha as informações de cliente, usuário e senha
        session.findById("wnd[0]/usr/txtRSYST-MANDT").text = "300"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = usuario
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = senha
        session.findById("wnd[0]").sendVKey(0)
    except Exception as e:
        print(f"Ocorreu um erro durante a autenticação no SAP: {e}")



# FUNÇÃO ENCERRAR O SAP
def close_process(nome_processo):
    for proc in psutil.process_iter(['pid', 'name']):
        if nome_processo.lower() in proc.info['name'].lower():
            try:
                processo = psutil.Process(proc.info['pid'])
                processo.terminate()  # ou processo.kill() para forçar o fechamento
                print(f'{proc.info["name"]} (PID: {proc.info["pid"]}) foi fechado.')
            except psutil.NoSuchProcess:
                print(f'Erro: Processo {proc.info["name"]} não encontrado.')
            except psutil.AccessDenied:
                print(f'Erro: Acesso negado para fechar {proc.info["name"]}.')


def colar_clipboard():
    text_codigos_material.insert(tk.END, pyperclip.paste() + "\n")



pasta = r'G:\Drives compartilhados\Custos_DC03\07_SAs\Análise de Saldo Automatizada'


# Obtém a data de hoje
hoje = datetime.now()

# Formata a data no formato DDMMYYYY
data_formatada = hoje.strftime('%d%m%Y')

print(data_formatada)


def executar_rotina():
    # Armazena informações
    # codigos_material = text_codigos_material.get("1.0", tk.END).strip()  # Obtém todos os códigos do material
    numero_ss = entry_ss.get()
    centro = entry_centro.get()
    obra = entry_obra.get()

    # Obter os códigos do material
    raw_codigos = text_codigos_material.get("1.0", tk.END).strip()
    
    # Converter a string em uma lista
    codigos_material = raw_codigos.split('\n')

    # Converter a lista para uma Series
    series = pd.Series(codigos_material)
    
    # Copiar para a área de transferência
    series.to_clipboard(index=False, header=False)
       
    print(f"Códigos de Material: {codigos_material}")
    print(f"Número da SS: {numero_ss}")
    print(f"Centro: {centro}")

    try:
        close_process("saplogon.exe")
        close_process("excel.exe")
    except:
        pass
        
    # Fazer login no SAP
    fazer_login()

    #INFORMAÇÕES SAP
    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "cn52n"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[1]/usr/ctxtTCNT-PROF_DB").text = "000000000001"
    session.findById("wnd[1]/usr/ctxtTCNT-PROF_DB").caretPosition = 12    
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtP_DISVAR").text = "/D CUSTOS"
    session.findById("wnd[0]").sendVKey(4)
    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[0,24]").text = "*"+centro+"*"
    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[0,24]").caretPosition = 8
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/usr/lbl[1,3]").caretPosition = 25
    session.findById("wnd[1]").sendVKey(2)
    session.findById("wnd[0]/usr/btn%_CN_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    session.findById("wnd[1]").sendVKey(8)
    session.findById("wnd[0]").sendVKey(8)
    session.findById("wnd[0]/usr/cntlALVCONTAINER/shellcont/shell").pressToolbarContextButton("&MB_EXPORT")
    session.findById("wnd[0]/usr/cntlALVCONTAINER/shellcont/shell").selectContextMenuItem("&XXL")
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = pasta
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "CN52N.xlsx"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 5
    session.findById("wnd[1]/tbar[0]/btn[11]").press()

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nmb51"
    session.findById("wnd[0]").sendVKey(0)

    # Copiar para a área de transferência
    series.to_clipboard(index=False, header=False)

    session.findById("wnd[0]/usr/radRFLAT_L").select()
    session.findById("wnd[0]/usr/ctxtALV_DEF").text = "/D CUSTOS"
    session.findById("wnd[0]/usr/btn%_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]").sendVKey(24)
    session.findById("wnd[1]").sendVKey(8)
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").text = obra
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtWERKS-LOW").caretPosition = 4
    session.findById("wnd[0]").sendVKey(8)
    session.findById("wnd[0]").sendVKey(16)
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = pasta
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "MB51.xlsx"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 4
    session.findById("wnd[1]").sendVKey(11)

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nme2l"
    session.findById("wnd[0]").sendVKey(0)
    
    # Copiar para a área de transferência
    series.to_clipboard(index=False, header=False)
    
    session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").text = obra
    session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").caretPosition = 4
    session.findById("wnd[0]/usr/btn%_S_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    session.findById("wnd[1]").sendVKey(8)
    session.findById("wnd[0]/usr/ctxtS_BEDAT-LOW").text = "01012020"
    session.findById("wnd[0]/usr/ctxtS_BEDAT-HIGH").text = data_formatada
    session.findById("wnd[0]/usr/ctxtS_BED  AT-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_BEDAT-LOW").caretPosition = 6
    session.findById("wnd[0]").sendVKey(8)
    
    session.findById("wnd[0]").sendVKey(33)
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").setCurrentCell(-1,"TEXT")
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").selectColumn("TEXT")
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").pressColumnHeader("TEXT")
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").setCurrentCell(1,"TEXT")
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").firstVisibleRow = 0
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").selectedRows = "1"
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").clickCurrentCell()
              
    session.findById("wnd[0]/tbar[1]/btn[43]").press()
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = pasta
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ME2L.xlsx"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 4
    session.findById("wnd[1]").sendVKey(11)

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nme5a"
        session.findById("wnd[0]").sendVKey(0)
    
        # Copiar para a área de transferência
        series.to_clipboard(index=False, header=False)
            
        session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").text = obra
        session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").setFocus
        session.findById("wnd[0]/usr/ctxtS_WERKS-LOW").caretPosition = 4
        session.findById("wnd[0]/usr/btn%_BA_MATNR_%_APP_%-VALU_PUSH").press()
        session.findById("wnd[1]").sendVKey(24)
        session.findById("wnd[1]").sendVKey(8)
        session.findById("wnd[0]").sendVKey(8)
        session.findById("wnd[0]/tbar[1]/btn[43]").press()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = pasta
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ME5A.xlsx"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 4
        session.findById("wnd[1]").sendVKey(11)

    except Exception as e:
        print(f"Erro ocorreu: {e}")
        
        # Criar um arquivo ME5A.xlsx vazio no caminho especificado
        arquivo_excel = os.path.join(pasta, "ME5A.xlsx")
        df = pd.DataFrame()  # Criar um DataFrame vazio
        df.to_excel(arquivo_excel, index=False)
        print(f"Arquivo {arquivo_excel} criado com sucesso.")
        
    # Fechar o SAP
    close_process("saplogon.exe")
    
    try:
        close_process("excel.exe")
    except:
        pass

    time.sleep(5)
       
    # Lista todos os arquivos na pasta com extensões .xlsx
    arquivos = [f for f in os.listdir(pasta) if f.lower().endswith('.xlsx')]
    
    # Verifica se há arquivos na pasta
    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado na pasta.")
    else:
        # Caminho do arquivo Excel final
        file_path = os.path.join(pasta, 'Analise SS - ' + numero_ss + '.xlsx')
    
        # Cria um objeto ExcelWriter para o arquivo final
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            abas_criadas = set()  # Para rastrear as abas já criadas
            aba_criada = False
            for arquivo in arquivos:
                caminho_arquivo = os.path.join(pasta, arquivo)
                # Verifica se o arquivo não está vazio e é um arquivo Excel válido
                if os.path.getsize(caminho_arquivo) > 0:
                    try:
                        # Lê cada arquivo Excel
                        df = pd.read_excel(caminho_arquivo, engine='openpyxl')
                        # Nome da aba será o nome do arquivo sem a extensão .xlsx
                        nome_aba = os.path.splitext(arquivo)[0]
    
                        # Verifica se a aba já foi criada
                        if nome_aba not in abas_criadas:
                            # Escreve o DataFrame na aba correspondente
                            df.to_excel(writer, sheet_name=nome_aba, index=False)
                            abas_criadas.add(nome_aba)  # Marca a aba como criada
                            aba_criada = True
                        else:
                            print(f"A aba '{nome_aba}' já foi criada, ignorando o arquivo {arquivo}.")
    
                    except Exception as e:
                        print(f"Erro ao ler o arquivo {arquivo}: {e}")
                else:
                    print(f"O arquivo {arquivo} está vazio e foi ignorado.")
    
            # Adiciona uma aba vazia se nenhuma aba foi criada
            if not aba_criada:
                pd.DataFrame({'Mensagem': ['Nenhum dado disponível']}).to_excel(writer, sheet_name='Vazio')
    
        print("Arquivos unificados com sucesso!")

                # Verifica se o arquivo foi criado corretamente
        if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
            print("Erro: O arquivo Excel não foi criado corretamente.")
        else:
            try:
                # Carrega o arquivo Excel
                workbook = load_workbook(file_path)
                # Adiciona uma nova aba chamada "Análise de Saldo"
                workbook.create_sheet(title="Análise de Saldo")
                # Salva o arquivo com a nova aba
                workbook.save(file_path)
                print("Aba 'Análise de Saldo' inserida com sucesso!")
            except Exception as e:
                print(f"Erro ao manipular o arquivo Excel: {e}")
                
        # Caminho do arquivo Excel
        file_path = os.path.join(pasta, 'Analise SS - ' + numero_ss + '.xlsx')
    
        # Carrega o arquivo Excel
        workbook = load_workbook(file_path)
    
        # Adiciona uma nova aba chamada "Análise de Saldo"
        workbook.create_sheet(title="Análise de Saldo")
    
        # Salva o arquivo com a nova aba
        workbook.save(file_path)
    
        print("Aba 'Análise de Saldo' inserida com sucesso!")
    
        # Carrega o arquivo Excel novamente para inserir valores
        workbook = load_workbook(file_path)
        sheet = workbook["Análise de Saldo"]
    
        # Insere os valores nas células
        sheet["B2"] = "DIAGRAMA"
        sheet["C2"] = "TAREFA"
        sheet["D2"] = "ITEM"
        sheet["E2"] = "MATERIAL"
        sheet["F2"] = "101"
        sheet["G2"] = "102"
        sheet["H2"] = "701"
        sheet["I2"] = "702"
        sheet["J2"] = "Z03"
        sheet["K2"] = "Somas"
        sheet["L2"] = "Subtrações"
        sheet["M2"] = "Recebido"
        sheet["N2"] = "Orçamento CN52N"
        sheet["O2"] = "Saldo não recebido"
        sheet["P2"] = "CN52N (LIB/CNPA)"
        sheet["Q2"] = "Saldo total"
        sheet["R2"] = "ME2L"
        sheet["S2"] = "Saldo em pedidos"
        sheet["T2"] = "Quantidade solicitada na SS/SA"
        sheet["U2"] = "Saldo ok?"
        sheet["V2"] = "Supressão de pedido?"
        sheet["W2"] = "Quantidade a suprimir"
        
        # Mescla as células C1:G1 e insere o nome "MB51"
        sheet.merge_cells("F1:J1")
        sheet["F1"] = "MB51"
    
        # Alinha o texto no centro para a célula mesclada
        sheet["F1"].alignment = Alignment(horizontal="center", vertical="center")
    
        # Salva as alterações no arquivo
        workbook.save(file_path)
    
        print("Aba 'Análise de Saldo' atualizada com sucesso!")
    
        # Variável com códigos de material
        codigos_material = raw_codigos.splitlines()
    
        # Insere os dados da variável codigos_material abaixo da célula B2
        start_row = 3  # Começa na célula B3
        for index, codigo in enumerate(codigos_material):
            sheet[f"E{start_row + index}"] = codigo
    
        # Salva as alterações no arquivo
        workbook.save(file_path)
    
        print("Dados da variável 'codigos_material' inseridos com sucesso na coluna 'Material'!")
    
        # Leia o arquivo Excel e a aba específica "CN52N"
        df = pd.read_excel(file_path, sheet_name="CN52N", engine='openpyxl')
        
        # Verifique se a coluna "Status do sistema" existe
        if "Status do sistema" in df.columns:
            # Substitua os valores na coluna "Status do sistema"
            df["Status do sistema"] = df["Status do sistema"].apply(lambda x: 1 if not str(x).startswith("BAIX") else x)
        else:
            print("A coluna 'Status do sistema' não foi encontrada no arquivo.")
    
        # Salve as alterações de volta no arquivo Excel
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name="CN52N", index=False)
    
        # Carrega o arquivo Excel novamente para inserir as fórmulas
        workbook = load_workbook(file_path)
        sheet = workbook["Análise de Saldo"]
    
        # Define a última linha da coluna 'Material' que contém dados
        max_row = sheet.max_row
    
        # Loop através de todas as linhas para verificar onde há códigos de material
        for row in range(3, max_row + 1):  # Começa na linha 3 e vai até a última linha
            if sheet[f"B{row}"].value:  # Verifica se há um código de material na coluna B
                # Inserir as fórmulas
                sheet[f"C{row}"] = f'=SUMIFS(\'MB51\'!$I:$I,\'MB51\'!$F:$F,C$2,\'MB51\'!$B:$B,$B{row})'
                sheet[f"D{row}"] = f'=SUMIFS(\'MB51\'!$I:$I,\'MB51\'!$F:$F,D$2,\'MB51\'!$B:$B,$B{row})'
                sheet[f"E{row}"] = f'=SUMIFS(\'MB51\'!$I:$I,\'MB51\'!$F:$F,E$2,\'MB51\'!$B:$B,$B{row})'
                sheet[f"F{row}"] = f'=SUMIFS(\'MB51\'!$I:$I,\'MB51\'!$F:$F,F$2,\'MB51\'!$B:$B,$B{row})'
                sheet[f"G{row}"] = f'=SUMIFS(\'MB51\'!$I:$I,\'MB51\'!$F:$F,G$2,\'MB51\'!$B:$B,$B{row})'
                sheet[f"H{row}"] = f'=C{row}+E{row}+G{row}'
                sheet[f"I{row}"] = f'=D{row}+F{row}'
                sheet[f"J{row}"] = f'=H{row}+I{row}'
                sheet[f"L{row}"] = f'=K{row}-J{row}'
                sheet[f"K{row}"] = f'=SUMIFS(\'CN52N\'!$J:$J,\'CN52N\'!$G:$G,$B{row})'
                sheet[f"M{row}"] = f'=SUMIFS(\'CN52N\'!$L:$L,\'CN52N\'!$I:$I,"<>BAIX*",\'CN52N\'!$G:$G,$B{row})'
                sheet[f"N{row}"] = f'=IF(L{row}<M{row},L{row},M{row})'            
                sheet[f"O{row}"] = f'=SUMIFS(\'ME2L\'!$H:$H,\'ME2L\'!$F:$F,$B{row})'                       
                sheet[f"P{row}"] = f'=K{row}-O{row}'
                sheet[f"S{row}"] = f'=IF(R{row}<=N{row},"OK","NÃO OK")'
                sheet[f"T{row}"] = f'=IF(OR(S{row}="NÃO OK",P{row}>=R{row}),"NÃO","SIM")'
                sheet[f"U{row}"] = f'=IF(T{row}="NÃO",0,O{row}-K{row}+R{row})'
    
        # Salva as alterações no arquivo
        workbook.save(file_path)
    
        print("Fórmulas inseridas com sucesso!")
    
        # Carregar o workbook e a planilha
        wb = load_workbook(file_path, data_only=False)
    
        # Selecionar a aba "Análise de Saldo"
        ws_main = wb['Análise de Saldo']
    
        # Remover linhas de grade da aba principal
        ws_main.sheet_view.showGridLines = False
    
        # Ajustar largura das colunas com base na segunda linha
        col_widths = {}
        for cell in ws_main[2]:  # Usar a segunda linha para ajustar a largura das colunas
            if cell.value:
                col_letter = cell.column_letter
                col_widths[col_letter] = max(col_widths.get(col_letter, 0), len(str(cell.value)))
    
        for col_letter, width in col_widths.items():
            ws_main.column_dimensions[col_letter].width = (width + 2) * 1.2  # Ajuste com margem de segurança
    
        # Congelar painéis: Primeira linha e primeira coluna
        ws_main.freeze_panes = "B3"
    
        # Salvar as alterações
        wb.save(file_path)
        wb.close()
    
        print("Personalizações aplicadas com sucesso!")

        # Listar todos os arquivos na pasta
        arquivos = os.listdir(pasta)
        
        # Loop sobre os arquivos e remover aqueles que não começam com "Analise SS"
        for arquivo in arquivos:
            if not arquivo.startswith("Analise SS"):
                caminho_completo = os.path.join(pasta, arquivo)
                if os.path.isfile(caminho_completo):  # Verifica se é um arquivo
                    os.remove(caminho_completo)
                    print(f'Removido: {caminho_completo}')

        # Caminho da pasta
        folder_path = r'G:\Drives compartilhados\Custos_DC03\07_SAs\Análise de Saldo Automatizada'
        
        # Lista de abas permitidas
        allowed_sheets = ["CN52N", "MB51", "ME2L", "ME5A", "Análise de Saldo"]
        
        # Percorrer todos os arquivos da pasta
        for filename in os.listdir(folder_path):
            # Verificar se o arquivo começa com "Analise SS" e é um arquivo Excel
            if filename.startswith("Analise SS") and filename.endswith(".xlsx"):
                file_path = os.path.join(folder_path, filename)
                
                # Carregar o workbook existente
                workbook = load_workbook(file_path)
                sheet_names = workbook.sheetnames
                
                # Verificar se existem abas não permitidas
                sheets_to_remove = [sheet for sheet in sheet_names if sheet not in allowed_sheets]
                
                if sheets_to_remove:
                    for sheet in sheets_to_remove:
                        # Remover a aba não permitida
                        workbook.remove(workbook[sheet])
                    
                    # Salvar o arquivo com as abas removidas
                    workbook.save(file_path)
                    print(f"Arquivo {filename} atualizado. Abas removidas: {sheets_to_remove}")
                else:
                    print(f"O arquivo {filename} já está com as abas corretas.")
 
    messagebox.showinfo("Execução", "Rotina executada com sucesso!")

# Criando a janela principal
root = tk.Tk()
root.title("Execução de Rotinas")
root.geometry("450x650")  # Ajustar o tamanho da janela
root.configure(bg='#f2f2f2')
root.resizable(False, False)  # Impede o redimensionamento da janela

# Fonte personalizada
root.option_add("*Font", "Amiko 10")

# Título centralizado
label_titulo = tk.Label(root, text="Gerador de Análise de SS", font=("Amiko", 16, "bold"), bg='#f2f2f2', fg='#b23a48')
label_titulo.pack(pady=(20, 10))

# Frame principal para centralizar os elementos
frame = tk.Frame(root, bg='#f2f2f2', padx=20, pady=10)
frame.pack(expand=True)

# Configuração de estilo para os widgets
input_style = {'font': ('Amiko', 10), 'bg': '#ffffff', 'bd': 0, 'highlightthickness': 1,
               'highlightbackground': '#d1d1d1', 'highlightcolor': '#b23a48', 'relief': 'flat'}

label_style = {'bg': '#f2f2f2', 'font': ('Amiko', 9)}
button_style = {'font': ('Amiko', 10), 'bg': '#b23a48', 'fg': 'white', 'bd': 0,
                'activebackground': '#8c2e39', 'relief': 'flat', 'cursor': 'hand2'}

# Label e input para o usuário
label_usuario = tk.Label(frame, text="Usuário", **label_style)
label_usuario.grid(row=0, column=0, padx=5, pady=(5, 0), sticky='w')
entry_usuario = tk.Entry(frame, **input_style, width=25)
entry_usuario.grid(row=1, column=0, columnspan=2, padx=5, pady=(0, 10))

# Label e input para a senha
label_senha = tk.Label(frame, text="Senha", **label_style)
label_senha.grid(row=2, column=0, padx=5, pady=(5, 0), sticky='w')
entry_senha = tk.Entry(frame, show="*", **input_style, width=25)
entry_senha.grid(row=3, column=0, columnspan=2, padx=5, pady=(0, 10))

# Label e input para o Centro
label_centro = tk.Label(frame, text="Obra (Nome)", **label_style)
label_centro.grid(row=4, column=0, padx=5, pady=(5, 0), sticky='w')
entry_centro = tk.Entry(frame, **input_style, width=15)
entry_centro.grid(row=5, column=0, padx=5, pady=(0, 10))

# Label e input para o Obra
label_obra = tk.Label(frame, text="Código Obra", **label_style)
label_obra.grid(row=4, column=1, padx=5, pady=(5, 0), sticky='w')
entry_obra = tk.Entry(frame, **input_style, width=15)
entry_obra.grid(row=5, column=1, padx=5, pady=(0, 10))

# Label e scrolled text para códigos de material
label_codigos = tk.Label(frame, text="Códigos de Material", **label_style)
label_codigos.grid(row=6, column=0, padx=5, pady=(5, 0), sticky='w')
text_codigos_material = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=30, height=4, font=('Amiko', 10), bg='#ffffff', bd=0,
                                                  relief='flat', highlightthickness=1, highlightbackground='#d1d1d1', highlightcolor='#b23a48')

text_codigos_material.grid(row=7, column=0, columnspan=2, padx=5, pady=(0, 10))

# Botão para colar do clipboard
button_colar = tk.Button(frame, text="Colar do Clipboard", command=colar_clipboard, **button_style, width=20)
button_colar.grid(row=8, column=0, columnspan=2, pady=(5, 10))

# Label e input para o número da SS
label_ss = tk.Label(frame, text="Número da SS", **label_style)
label_ss.grid(row=9, column=0, padx=5, pady=(5, 0), sticky='w')
entry_ss = tk.Entry(frame, **input_style, width=25)
entry_ss.grid(row=10, column=0, columnspan=2, padx=5, pady=(0, 10))

# Botão para executar
button_executar = tk.Button(frame, text="Executar", command=executar_rotina, **button_style, width=20)
button_executar.grid(row=11, column=0, columnspan=2, pady=(10, 20))

# Configurar o comportamento de hover para os botões
def on_enter(e):
    e.widget['bg'] = '#8c2e39'

def on_leave(e):
    e.widget['bg'] = '#b23a48'

button_colar.bind("<Enter>", on_enter)
button_colar.bind("<Leave>", on_leave)
button_executar.bind("<Enter>", on_enter)
button_executar.bind("<Leave>", on_leave)

# Inicializa a interface
root.mainloop()

