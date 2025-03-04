# Bibliotecas da biblioteca padrão
import time
import subprocess
import os
import shutil
import psutil
import glob
import shlex
from datetime import datetime, timedelta
from io import StringIO
from tkinter import messagebox
import tkinter as tk
import warnings

# Bibliotecas de terceiros
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
import win32com.client
from dateutil.relativedelta import relativedelta
import webbrowser

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import base64
import pickle
import gspread
from google.auth.transport.requests import Request
from google.auth.exceptions import GoogleAuthError
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload


def nf_recebida():
    # 🔹 Token serializado embutido (cole o Base64 aqui)
    EMBEDDED_TOKEN_B64 = "gASVLwoAAAAAAACMGWdvb2dsZS5vYXV0aDIuY3JlZGVudGlhbHOUjAtDcmVkZW50aWFsc5STlCmBlH2UKIwFdG9rZW6UjOF5YTI5LmEwQVhlTzgwU0NnV2drNURFOGppc2RkM1dpb3VWR0ExMm9MZ0NQTTBBSWRJNlJKX05LYkUtNVlOaUI4S2JkdEdjU1VCSkl3Q2s2c3JsZ1M5bHdVUjdFQmdEb1M0dVVtb0E4a0NJN01OcnkySHpwcEJsS29HN0ZiVDR4U1pxQUU2bzRqVFhiTlIzU1NDMHN3VHJsYzNhZ09pRFZSRjFFV3NpSEhKeDM4ZmhsZ25nYUNnWUtBVU1TQVJNU0ZRSEdYMk1peTJZSmFSeS01dy1jODhBdi1YQmYzUTAxNziUjAZleHBpcnmUjAhkYXRldGltZZSMCGRhdGV0aW1llJOUQwoH6QIMDgIdDojulIWUUpSMDl9yZWZyZXNoX3Rva2VulIxnMS8vMGh1Ry1VSEJCWmpFOENnWUlBUkFBR0JFU053Ri1MOUlySk12bllaN1NidVNBelNUVWpad2Vtc0Nvcm50ZlVmaFVxdllOX3g4QTI0QzJBOUZjenpzOGJ2Vms1QkgzT3ZTdURISZSMCV9pZF90b2tlbpRYwQQAAGV5SmhiR2NpT2lKU1V6STFOaUlzSW10cFpDSTZJbVZsWXpVek5HWmhOV0k0WTJGallUSXdNV05oT0dRd1ptWTVObUkxTkdNMU5qSXlNVEJrTVdVaUxDSjBlWEFpT2lKS1YxUWlmUS5leUpwYzNNaU9pSm9kSFJ3Y3pvdkwyRmpZMjkxYm5SekxtZHZiMmRzWlM1amIyMGlMQ0poZW5BaU9pSXpPVEF6TVRZME56WXhNVEl0YUhOdFpYRnFhV3B5ZEd4cE9ERnROVFYwTW1OcmEycHBaMlkzWnpBd2RXUXVZWEJ3Y3k1bmIyOW5iR1YxYzJWeVkyOXVkR1Z1ZEM1amIyMGlMQ0poZFdRaU9pSXpPVEF6TVRZME56WXhNVEl0YUhOdFpYRnFhV3B5ZEd4cE9ERnROVFYwTW1OcmEycHBaMlkzWnpBd2RXUXVZWEJ3Y3k1bmIyOW5iR1YxYzJWeVkyOXVkR1Z1ZEM1amIyMGlMQ0p6ZFdJaU9pSXhNRE01TlRJMk16TTRNekU0TlRRek16VXhPREFpTENKb1pDSTZJbkp2WjJkaExtTnZiUzVpY2lJc0ltVnRZV2xzSWpvaWJIVmpZWE11WjJGc2FXTnBiMnhwUUhKdloyZGhMbU52YlM1aWNpSXNJbVZ0WVdsc1gzWmxjbWxtYVdWa0lqcDBjblZsTENKaGRGOW9ZWE5vSWpvaVlXeGhPV1pPWTA5ck5FSnFiVloyVjJkeE9FUm9keUlzSW01aGJXVWlPaUpNZFdOaGN5Qk5ZWEpqWld3Z1IyRnNhV05wYjJ4cElpd2ljR2xqZEhWeVpTSTZJbWgwZEhCek9pOHZiR2d6TG1kdmIyZHNaWFZ6WlhKamIyNTBaVzUwTG1OdmJTOWhMMEZEWnpodlkwcFhUSGhZWkhCYVNsTlRPRkYzVEhKS1pGZzFNek5VZUVFeExUWmtSRUV0Vm05NVYzTjZWMHQ1UjJ3MFJGQlZZVUpXTUVFOWN6azJMV01pTENKbmFYWmxibDl1WVcxbElqb2lUSFZqWVhNaUxDSm1ZVzFwYkhsZmJtRnRaU0k2SWsxaGNtTmxiQ0JIWVd4cFkybHZiR2tpTENKcFlYUWlPakUzTXprek5qVXpOVEFzSW1WNGNDSTZNVGN6T1RNMk9EazFNSDAuaC16N085YWViQzRNcGNlaWl6Z1hoVFRwaUVadkJwNC1yZXp6SjlMUTFTV0dnSldlNndZOUEwY0sxUkcxbkVVNlFNTlBUMC1mSG5KTzhic2VJWVdzNjZzLVhoSEh0S0dfZUw4U2lETjViTEVSV25BTERmODdROElIUnlaYXZPajJ0MXlnNTBnVFIxVkQ0bzdOdGdKa2xGXzd1Wnl2UTdYZjc2Z2JfWmxVdlBYcEZlZFRfUmg4OWlZRl9oOUZIWjZyLUlQVlpROE1iNlNHZ0lFN0cxU1J0WXdTelRYOThsTk51Q3JiclB3NmRiNGNBbHJaSGN6R0Q1Z3YzaC14TUZUbFVvUks0NHdHOEZzSDVRVnQzaVBaUC1Yb1VRSFZGeFV5al9yWktRMlV5dFJOSVFlNEc1QnZkNVMxVGZ6RDgwN3dUQl9JOXJWbWU3NzZyWC1FeWtiWmpnlIwHX3Njb3Blc5RdlCiMBm9wZW5pZJSMMGh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvdXNlcmluZm8ucHJvZmlsZZSMLmh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvdXNlcmluZm8uZW1haWyUjCxodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL3NwcmVhZHNoZWV0c5SMJWh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvZHJpdmWUjCpodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL2RyaXZlLmZpbGWUZYwPX2RlZmF1bHRfc2NvcGVzlE6MD19ncmFudGVkX3Njb3Blc5RdlCiMKmh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvZHJpdmUuZmlsZZSMLmh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvdXNlcmluZm8uZW1haWyUjDBodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL3VzZXJpbmZvLnByb2ZpbGWUjCxodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL3NwcmVhZHNoZWV0c5SMBm9wZW5pZJSMJWh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvZHJpdmWUZYwKX3Rva2VuX3VyaZSMI2h0dHBzOi8vb2F1dGgyLmdvb2dsZWFwaXMuY29tL3Rva2VulIwKX2NsaWVudF9pZJSMSDM5MDMxNjQ3NjExMi1oc21lcWppanJ0bGk4MW01NXQyY2tramlnZjdnMDB1ZC5hcHBzLmdvb2dsZXVzZXJjb250ZW50LmNvbZSMDl9jbGllbnRfc2VjcmV0lIwjR09DU1BYLTRsT2tMYmtMR1pUZzg1eHYyMU5pTUNtZTdsVWGUjBFfcXVvdGFfcHJvamVjdF9pZJROjAtfcmFwdF90b2tlbpROjBZfZW5hYmxlX3JlYXV0aF9yZWZyZXNolImMD190cnVzdF9ib3VuZGFyeZROjBBfdW5pdmVyc2VfZG9tYWlulIwOZ29vZ2xlYXBpcy5jb22UjA9fY3JlZF9maWxlX3BhdGiUTowZX3VzZV9ub25fYmxvY2tpbmdfcmVmcmVzaJSJjAhfYWNjb3VudJSMAJR1Yi4="

    # 🔹 Ordem correta das colunas conforme Google Sheets
    google_columns = [
        "Status",
        "Responsável - Etapa Atual",
        "Observação",
        "Local",
        "Disponível para lançamento?",
        "Área",
        "E-mail",
        "Concluído?",
        "Data Atualização",
        "Nº NF-e",
        "Chave de acesso",
        "Data processamento",
        "CNPJ do emissor",
        "CNPJ destinatário",
        "Nome emissor",
        "Val.total c/impostos",
        "Data Vencimento",
        "Status global",
        "Local de negócios",
    ]

    # 🔹 Decodifica o token embutido e carrega em memória
    try:
        token_bytes = base64.b64decode(EMBEDDED_TOKEN_B64)
        creds = pickle.loads(token_bytes)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        print("✅ Credenciais carregadas com sucesso!")
    except (pickle.PickleError, GoogleAuthError, base64.binascii.Error) as e:
        print(f"❌ Erro ao carregar as credenciais embutidas: {e}")
        creds = None

    if not creds:
        raise Exception("❌ Falha ao carregar as credenciais embutidas!")

    # 🔹 Conectar ao Google Sheets e Drive
    client = gspread.authorize(creds)
    drive_service = build("drive", "v3", credentials=creds)

    SPREADSHEET_ID = "15gkc4fpd1vzp6cAOwmtYgXkhYOODUCqDgHOr5_8DWPE"
    SHEET_NAME = "Notas_Fiscais"
    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

    # 🔹 Carregar os dados existentes do Google Sheets
    existing_data = sheet.get_all_values()

    # 🔹 Identificar o índice correto das colunas no Google Sheets
    header = [col.strip() for col in existing_data[0]]
    col_index = {col: header.index(col) for col in google_columns if col in header}

    if "Chave de acesso" not in col_index:
        raise Exception(
            "❌ A coluna 'Chave de acesso' não foi encontrada no Google Sheets!"
        )

    existing_keys = {
        row[col_index["Chave de acesso"]]: i + 2
        for i, row in enumerate(existing_data[1:])
        if len(row) > col_index["Chave de acesso"]
        and row[col_index["Chave de acesso"]].strip()
    }

    update_columns = [
        "Status",
        "Responsável - Etapa Atual",
        "Observação",
        "Local",
        "Disponível para lançamento?",
        "Área",
        "E-mail",
        "Concluído?",
        "Data Atualização",
        "Nº NF-e",
        "Chave de acesso",
        "Data processamento",
        "CNPJ do emissor",
        "CNPJ destinatário",
        "Nome emissor",
        "Val.total c/impostos",
        "Data Vencimento",
        "Status global",
        "Local de negócios",
    ]
    update_indices = [col_index[col] for col in update_columns]

    FILES = [
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    ]

    def upload_excel_to_gsheets(local_file_path):
        file_metadata = {
            "name": "Temp_Converted",
            "mimeType": "application/vnd.google-apps.spreadsheet",
        }
        media = MediaFileUpload(
            local_file_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file = (
            drive_service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )
        print(f"✅ Arquivo convertido para Google Sheets. ID: {file['id']}")
        return file["id"]

    new_rows = []
    updates = []
    for file in FILES:
        google_sheet_id = upload_excel_to_gsheets(file)
        time.sleep(3)
        temp_sheet = client.open_by_key(google_sheet_id).worksheet("Célula Fiscal")

        df = pd.DataFrame(temp_sheet.get_all_records())
        df.columns = [str(col).strip() for col in df.columns]
        print(f"✅ Nomes das colunas encontrados no arquivo: {df.columns.tolist()}")

        if "Chave de acesso" not in df.columns:
            raise ValueError(f"⚠️ O arquivo não contém a coluna 'Chave de acesso'.")

        for _, row in df.iterrows():
            chave = str(row.get("Chave de acesso", "")).strip()
            valores_atualizados = [
                str(row.get(col, "")).strip() if pd.notna(row.get(col, "")) else ""
                for col in update_columns
            ]

            if chave and chave in existing_keys:
                linha_google = existing_keys[chave]
                col_start = "E"
                col_end = chr(ord(col_start) + len(update_columns) - 1)
                updates.append(
                    {
                        "range": f"{col_start}{linha_google}:{col_end}{linha_google}",
                        "values": [valores_atualizados],
                    }
                )
            elif chave:
                row_dict = {col: row.get(col, "") for col in google_columns}
                formatted_row = [
                    str(row_dict[col]) if pd.notna(row_dict[col]) else ""
                    for col in google_columns
                ]
                new_rows.append(formatted_row)

        drive_service.files().delete(fileId=google_sheet_id).execute()
        print("🗑️ Arquivo temporário removido do Google Drive.")

    if new_rows:
        start_col = 5
        formatted_rows = [[""] * (start_col - 1) + row for row in new_rows]
        sheet.append_rows(formatted_rows)
        print(f"✅ {len(new_rows)} novas linhas adicionadas com sucesso!")

    if updates:
        sheet.batch_update(updates)
        print(f"✅ {len(updates)} registros atualizados com sucesso!")
    else:
        print(
            "✅ Nenhuma nova linha adicionada ou atualizada. Todos os dados já estavam sincronizados."
        )


def salvar_copia_celula_fiscal():
    # Suprimir warnings
    warnings.filterwarnings("ignore")

    # Definir caminho do arquivo original e o da pasta de destino
    caminho_arquivo_original = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    pasta_destino = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\Historico"

    # Definir a aba a ser copiada
    aba = "Célula Fiscal"

    # Ler a aba "Célula Fiscal" do arquivo original
    df = pd.read_excel(caminho_arquivo_original, sheet_name=aba)

    # Obter a data atual no formato DD.MM.AA
    data_atual = datetime.now().strftime("%d.%m.%y")

    # Iniciar o número da versão (incremental) como 1
    numero_versao = 1

    # Gerar o nome inicial do arquivo de log
    nome_arquivo_log = f"Log_CF_{data_atual}_v{numero_versao}.xlsx"
    caminho_arquivo_log = os.path.join(pasta_destino, nome_arquivo_log)

    # Verificar se já existe um arquivo com o mesmo nome e incrementar a versão
    while os.path.exists(caminho_arquivo_log):
        numero_versao += 1
        nome_arquivo_log = f"Log_CF_{data_atual}_v{numero_versao}.xlsx"
        caminho_arquivo_log = os.path.join(pasta_destino, nome_arquivo_log)

    # Salvar a cópia da aba "Célula Fiscal" no novo arquivo
    with pd.ExcelWriter(caminho_arquivo_log, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=aba)

    print(f"Arquivo salvo com sucesso: {nome_arquivo_log}")


def compilar_historico():
    # Caminho da pasta contendo os arquivos de log
    pasta_historico = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\Historico"

    # Caminho do arquivo compilado
    caminho_arquivo_compilado = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Historico_Compilado.xlsx"

    # Lista para armazenar os DataFrames
    lista_df = []

    # Percorrer todos os arquivos na pasta
    for arquivo in os.listdir(pasta_historico):
        if arquivo.endswith(".xlsx"):
            caminho_arquivo = os.path.join(pasta_historico, arquivo)

            # Extrair a data do nome do arquivo (assumindo que o formato é Log_CF_DD.MM.AA_vN.xlsx)
            partes_nome = arquivo.split("_")
            data_str = partes_nome[2]  # DD.MM.AA
            data_formatada = datetime.strptime(data_str, "%d.%m.%y").strftime(
                "%d/%m/%Y"
            )  # Formatar como DD/MM/AAAA

            # Ler a aba "Célula Fiscal" de cada arquivo
            df = pd.read_excel(caminho_arquivo, sheet_name="Célula Fiscal")

            # Excluir colunas indesejadas
            colunas_excluir = [
                "Local",
                "Disponível para lançamento?",
                "Área",
                "E-mail",
                "Área ",
                "Data Atualização",
            ]
            df = df.drop(columns=colunas_excluir, errors="ignore")

            # Excluir linhas onde a célula da coluna 'Status' está em branco
            df = df.dropna(subset=["Status"])

            # Reordenar colunas - mover 'Nº NF-e' para a segunda posição e 'Concluído?' para a terceira
            colunas = df.columns.tolist()  # Obter todas as colunas
            if "Nº NF-e" in colunas and "Concluído?" in colunas:
                colunas.remove("Nº NF-e")
                colunas.remove("Concluído?")
                colunas.insert(1, "Nº NF-e")  # Inserir na segunda posição
                colunas.insert(2, "Concluído?")  # Inserir na terceira posição
                df = df[colunas]  # Reordenar DataFrame

            # Inserir a coluna "Data de Cobrança" como a primeira coluna
            df.insert(0, "Data de Cobrança", data_formatada)

            # Adicionar o DataFrame à lista
            lista_df.append(df)

    # Concatenar todos os DataFrames em um único DataFrame
    df_compilado = pd.concat(lista_df, ignore_index=True)

    # Salvar o DataFrame compilado em um arquivo Excel
    with pd.ExcelWriter(caminho_arquivo_compilado, engine="openpyxl") as writer:
        df_compilado.to_excel(writer, index=False, sheet_name="Historico")

        # Ajustar a largura das colunas e congelar a primeira linha
        worksheet = writer.sheets["Historico"]

        # Congelar a primeira linha
        worksheet.freeze_panes = worksheet[
            "B2"
        ]  # Congela após a primeira linha e a coluna A

        # Ajustar automaticamente a largura das colunas
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Coluna atual
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

    print(
        f"Arquivo de histórico compilado salvo com sucesso: {caminho_arquivo_compilado}"
    )


# FUNÇÃO ENCERRAR O SAP
def close_process(nome_processo):
    for proc in psutil.process_iter(["pid", "name"]):
        if nome_processo.lower() in proc.info["name"].lower():
            try:
                processo = psutil.Process(proc.info["pid"])
                processo.terminate()  # ou processo.kill() para forçar o fechamento
                print(f'{proc.info["name"]} (PID: {proc.info["pid"]}) foi fechado.')
            except psutil.NoSuchProcess:
                print(f'Erro: Processo {proc.info["name"]} não encontrado.')
            except psutil.AccessDenied:
                print(f'Erro: Acesso negado para fechar {proc.info["name"]}.')


def formulas_cf():
    caminho_arquivo = r"G:\\Drives compartilhados\\Fiscal_Arquivo_de_Notas\\FISCAL_MATERIAL\\CONTROLE GRC - MATERIAIS\\Célula Fiscal GRC - Materiais.XLSX"

    # Carregar a pasta de trabalho usando openpyxl
    workbook = openpyxl.load_workbook(caminho_arquivo)

    # Selecionar a aba "Célula Fiscal"
    sheet = workbook["Célula Fiscal"]

    # Remover o autofiltro, se existir
    if sheet.auto_filter.ref is not None:
        sheet.auto_filter.ref = None

    # Salvar o arquivo
    workbook.save(caminho_arquivo)

    # Define a última linha da coluna 'Material' que contém dados
    max_row = sheet.max_row

    # Loop através de todas as linhas para verificar onde há códigos de material
    for row in range(2, max_row + 1):  # Começa na linha 3 e vai até a última linha
        if sheet[f"L{row}"].value:  # Verifica se há um código de material na coluna B
            # Inserir as fórmulas
            sheet[f"D{row}"] = f"=IFERROR(VLOOKUP(S{row},'Dados'!A:C,2,0),)"
            sheet[f"E{row}"] = (
                f'=IF(R{row}="","",IF(R{row}=1,"Sim","Em Processamento"))'
            )
            sheet[f"F{row}"] = f"=IFERROR(VLOOKUP(B{row},'Dados'!C:D,2,0),)"
            sheet[f"G{row}"] = (
                f"=IFERROR(IF(OR(B{row}=\"Obra\",E{row}=\"Sim\"),VLOOKUP(D{row},'Dados'!B:E,4,0),VLOOKUP(B{row},'Dados'!C:E,3,0)),)"
            )

    # Salva as alterações no arquivo
    workbook.save(caminho_arquivo)


# Obter a data atual
data_atual = datetime.now()

# Calcular o primeiro dia do ano atual
primeiro_dia_ano = data_atual.replace(month=1, day=1)

# Converter a data para o formato desejado (ddmmaaaa)
M1 = primeiro_dia_ano.strftime("%d%m%Y")
MA = data_atual.strftime("%d%m%Y")

# Obter o ano corrente
ano_corrente = data_atual.year

# Calcular o primeiro dia de dois meses antes do mês atual
data_dois_meses_antes = data_atual - relativedelta(months=2)
primeiro_dia_dois_meses_antes = data_dois_meses_antes.replace(day=1)

# Converter para o formato desejado (ddmmaaaa)
M0 = primeiro_dia_dois_meses_antes.strftime("%d%m%Y")

# Exibir os resultados
print("Primeiro dia do ano atual (M1):", M1)
print("Data atual (MA):", MA)
print("Ano corrente:", ano_corrente)
print("Primeiro dia de dois meses atrás (M0):", M0)


def read_txt(file_path):
    def read_with_encoding(encoding):
        valid_lines = []
        with open(file_path, "r", encoding=encoding, errors="ignore") as file:
            for line in file:
                if line.count("|") == 35:  # Ajuste este valor conforme necessário
                    valid_lines.append(line)
        return valid_lines

    # Tenta ler o arquivo primeiro em UTF-8, depois em Latin-1 se falhar
    try:
        valid_lines = read_with_encoding("utf-8")
        if not valid_lines:
            raise UnicodeDecodeError
    except UnicodeDecodeError:
        valid_lines = read_with_encoding("latin-1")

    data = StringIO("".join(valid_lines))
    df = pd.read_csv(data, delimiter="|")
    return df


def save_to_excel(df, output_path):
    df.to_excel(output_path, index=False, engine="xlsxwriter")


def process_files(input_dir, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    all_files = [f for f in os.listdir(input_dir) if f.endswith(".txt")]

    for file in all_files:
        file_path = os.path.join(input_dir, file)
        df = read_txt(file_path)

        # Manter o mesmo nome do arquivo txt, apenas alterando a extensão para .xlsx
        output_file = os.path.join(output_dir, f"{os.path.splitext(file)[0]}.xlsx")
        save_to_excel(df, output_file)


def verificar_credenciais(email, senha):
    try:
        # Conectar ao servidor SMTP do Gmail
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login(email, senha)
        server.quit()
        return True
    except smtplib.SMTPAuthenticationError:
        return False
    except Exception as e:
        print(f"Erro: {e}")
        return False


def credenciais_login_email():
    usuario_gmail = entry_usuario_gmail.get()
    senha_gmail = entry_senha_gmail.get()
    return usuario_gmail, senha_gmail


def fazer_login():

    caminho_executavel_sap = r"C:\Program Files\SAP\FrontEnd\SAPGUI\saplogon.exe"
    if not os.path.exists(caminho_executavel_sap):
        caminho_executavel_sap = (
            r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        )

    print(f"Caminho do executável: {caminho_executavel_sap}")

    if not os.path.exists(caminho_executavel_sap):
        print("O arquivo não foi encontrado. Verifique o caminho.")
        messagebox.showerror(
            "Erro", "O executável do SAP GUI não foi encontrado. Verifique o caminho."
        )
        return

    # Use shlex.split para tratar corretamente o caminho com espaços
    comando = shlex.split(f'"{caminho_executavel_sap}"')
    print(f"Comando para execução: {comando}")

    try:
        subprocess.Popen(comando)
    except Exception as e:
        print(f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        return

    time.sleep(3)

    usuario_sap = entry_usuario_sap.get()
    senha_sap = entry_senha_sap.get()

    # Conexão com o SAP GUI
    try:
        sapguiauto = win32com.client.GetObject("SAPGUI")
        application = sapguiauto.GetScriptingEngine
        connection = application.OpenConnection("NOVO PVR - Produção SAP ECC", True)
    except Exception as e:
        print(
            f"Conexão 'NOVO PVR - Produção SAP ECC' não encontrada. Tentando outra conexão..."
        )
        try:
            connection = application.OpenConnection(
                "PVR - Produção SAP ECC (Interno)", True
            )
        except Exception as e:
            print(f"Ocorreu um erro ao abrir a conexão: {e}")
            close_process("saplogon.exe")
            return

    session = connection.Children(0)

    try:
        # Preencher as informações de cliente, usuário e senha
        session.findById("wnd[0]/usr/txtRSYST-MANDT").text = "300"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = usuario_sap
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = senha_sap
        session.findById("wnd[0]/usr/txtRSYST-LANGU").text = "PT"
        session.findById("wnd[0]").sendVKey(0)

    except:
        messagebox.showerror("Erro", f"Usuário ou senha incorretos")
        close_process("saplogon.exe")
        return


# Obter o diretório do usuário atual
user_dir = os.path.expanduser("~")

# Criar o caminho dinâmico para a pasta SAP\SAP GUI
sap_gui_path = os.path.join(user_dir, "Documents", "SAP", "SAP GUI")

print(sap_gui_path)


def excluir_arquivos_sap_gui():
    # Obtém todos os arquivos na pasta sap_gui_path
    arquivos = glob.glob(os.path.join(sap_gui_path, "*"))

    # Exclui cada arquivo encontrado
    for arquivo in arquivos:
        if os.path.isfile(arquivo):
            os.remove(arquivo)
            print(f"Arquivo {arquivo} excluído.")


def processar_arquivos(pasta_origem, pasta_destino):
    # Verifica se as pastas existem
    if not os.path.exists(pasta_origem):
        print(f"A pasta de origem '{pasta_origem}' não foi encontrada.")
        return

    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        print(f"A pasta de destino '{pasta_destino}' não existia, então foi criada.")

    # Itera sobre os arquivos na pasta de origem
    for arquivo in os.listdir(pasta_origem):
        caminho_arquivo_origem = os.path.join(pasta_origem, arquivo)

        # Verifica se é um arquivo
        if os.path.isfile(caminho_arquivo_origem):
            caminho_arquivo_destino = os.path.join(pasta_destino, arquivo)

            # Se o arquivo já existir no destino, remove o arquivo antigo
            if os.path.exists(caminho_arquivo_destino):
                os.remove(caminho_arquivo_destino)
                print(f"Arquivo {arquivo} existente no destino removido.")

            # Move o arquivo para a pasta de destino
            shutil.move(caminho_arquivo_origem, caminho_arquivo_destino)
            print(f"Arquivo {arquivo} movido para {pasta_destino}")


def criar_copia_temporaria(caminho_arquivo, caminho_temp):
    try:
        wb_original = load_workbook(caminho_arquivo, data_only=True)
        ws_original = wb_original["Célula Fiscal"]

        wb_temp = Workbook()
        ws_temp = wb_temp.active
        ws_temp.title = "Célula Fiscal"

        for row in ws_original.iter_rows(values_only=True):
            ws_temp.append(row)

        wb_temp.save(caminho_temp)
        print("Cópia temporária criada com sucesso.")
    except Exception as e:
        print(f"Erro ao criar a cópia temporária: {e}")
        exit()

    # Carregar a planilha temporária sem fórmulas
    try:
        df = pd.read_excel(caminho_temp, sheet_name="Célula Fiscal")
        print("Planilha carregada com sucesso.")
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        exit()


def GRC_ETL():
    excluir_arquivos_sap_gui()
    close_process("saplogon.exe")
    # Fazer login no SAP
    fazer_login()
    # INFORMAÇÕES SAP
    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "zbrinb006"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtS_STAT-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_STAT-LOW").caretPosition = 0
    session.findById("wnd[0]/usr/btn%_S_STAT_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/usr/tabsTAB_STRIP/tabpNOSV").select()
    session.findById(
        "wnd[1]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,0]"
    ).text = "99"
    session.findById(
        "wnd[1]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,1]"
    ).text = "98"
    session.findById(
        "wnd[1]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,2]"
    ).text = "89"
    session.findById(
        "wnd[1]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,2]"
    ).setFocus
    session.findById(
        "wnd[1]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,2]"
    ).caretPosition = 2
    session.findById("wnd[1]").sendVKey(8)

    # MATERIAL - Geral
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").text = "NORMPRCH"
    session.findById("wnd[0]/usr/ctxtS_CREDAT-LOW").text = "01072024"
    session.findById("wnd[0]/usr/ctxtS_CREDAT-HIGH").text = MA
    session.findById("wnd[0]/usr/ctxtS_CREDAT-HIGH").setFocus
    session.findById("wnd[0]/usr/ctxtS_CREDAT-HIGH").caretPosition = 6
    session.findById("wnd[0]").sendVKey(8)
    session.findById(
        "wnd[0]/usr/cntlCABECALHO/shellcont/shell"
    ).pressToolbarContextButton("&MB_EXPORT")
    session.findById("wnd[0]/usr/cntlCABECALHO/shellcont/shell").selectContextMenuItem(
        "&XXL"
    )
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZBRINB006 Material.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 18
    session.findById("wnd[1]").sendVKey(11)
    session.findById("wnd[0]").sendVKey(3)

    # Exemplo de uso da função
    pasta_origem_input = sap_gui_path
    pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - GERAL"

    processar_arquivos(pasta_origem_input, pasta_destino_input)

    # SERVIÇOS - Geral
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").caretPosition = 8
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").text = "ZSRVPRCH"
    session.findById("wnd[0]").sendVKey(8)
    session.findById(
        "wnd[0]/usr/cntlCABECALHO/shellcont/shell"
    ).pressToolbarContextButton("&MB_EXPORT")
    session.findById("wnd[0]/usr/cntlCABECALHO/shellcont/shell").selectContextMenuItem(
        "&XXL"
    )
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZBRINB006 Serviço.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 18
    session.findById("wnd[1]").sendVKey(11)
    session.findById("wnd[0]").sendVKey(3)

    pasta_origem_input = sap_gui_path
    pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - GERAL"

    processar_arquivos(pasta_origem_input, pasta_destino_input)

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nzbrinb006"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/radP_ERR").select()
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").text = "NORMPRCH"
    session.findById("wnd[0]/usr/ctxtS_CREDAT-LOW").text = M0
    session.findById("wnd[0]/usr/ctxtS_CREDAT-HIGH").text = MA
    session.findById("wnd[0]/usr/ctxtS_TPMSG-LOW").text = "E"
    session.findById("wnd[0]/usr/ctxtS_TPMSG-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_TPMSG-LOW").caretPosition = 1
    session.findById("wnd[0]").sendVKey(8)

    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").setCurrentCell(
        13, "BRANCH"
    )
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectedRows = "13"
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectContextMenuItem(
        "&XXL"
    )
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZBRINB006 Material.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 23
    session.findById("wnd[1]").sendVKey(11)
    session.findById("wnd[0]").sendVKey(3)

    pasta_origem_input = sap_gui_path
    pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - ERROS"

    processar_arquivos(pasta_origem_input, pasta_destino_input)

    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").text = "ZSRVPRCH"
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtS_PRCTYP-LOW").caretPosition = 8
    session.findById("wnd[0]").sendVKey(8)

    try:
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").setCurrentCell(
            16, "BUKRS"
        )
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectedRows = "16"
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").contextMenu()
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectContextMenuItem(
            "&XXL"
        )
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZBRINB006 Serviço.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 18
        session.findById("wnd[1]").sendVKey(11)

        pasta_origem_input = sap_gui_path
        pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - ERROS"

        processar_arquivos(pasta_origem_input, pasta_destino_input)

    except:
        pass

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nj1bnfe"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtDATE0-LOW").text = M1
    session.findById("wnd[0]/usr/ctxtDATE0-HIGH").text = MA
    session.findById("wnd[0]/usr/ctxtBUKRS-LOW").text = "1000"
    session.findById("wnd[0]/usr/ctxtBUKRS-LOW").setFocus
    session.findById("wnd[0]/usr/ctxtBUKRS-LOW").caretPosition = 4
    session.findById("wnd[0]").sendVKey(8)

    try:
        session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").setCurrentCell(
            11, "AUTHCOD"
        )
        session.findById("wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell").contextMenu()
        session.findById(
            "wnd[0]/usr/cntlNFE_CONTAINER/shellcont/shell"
        ).selectContextMenuItem("&XXL")
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "J1BNFE.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 6
        session.findById("wnd[1]").sendVKey(11)

        pasta_origem_input = sap_gui_path
        pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\J1BNFE"

        processar_arquivos(pasta_origem_input, pasta_destino_input)
    except:
        pass

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nmir5"
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]/usr/ctxtSO_BLDAT-LOW").text = M1
        session.findById("wnd[0]/usr/ctxtSO_BLDAT-HIGH").text = MA
        session.findById("wnd[0]/usr/ctxtSO_BUDAT-HIGH").setFocus
        session.findById("wnd[0]/usr/ctxtSO_BUDAT-HIGH").caretPosition = 6
        session.findById("wnd[0]").sendVKey(8)
        session.findById("wnd[0]").sendVKey(43)
        session.findById("wnd[1]").sendVKey(0)
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "MIR5.XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 4
        session.findById("wnd[1]").sendVKey(11)

        pasta_origem_input = sap_gui_path
        pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\MIR5"

        processar_arquivos(pasta_origem_input, pasta_destino_input)
    except:
        pass

    close_process("saplogon.exe")

    # Caminho do arquivo Excel
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"

    # Carregar a pasta de trabalho
    workbook = openpyxl.load_workbook(file_path)

    # Selecionar a aba "Célula Fiscal"
    sheet = workbook["Célula Fiscal"]

    # Remover o autofiltro, se existir
    if sheet.auto_filter.ref is not None:
        sheet.auto_filter.ref = None

    # Salvar o arquivo
    workbook.save(file_path)

    # Caminho do arquivo de origem
    origem = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - GERAL\ZBRINB006 Material.XLSX"

    # Caminho da pasta de destino
    destino = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\ZBRINB006 Material.XLSX"

    # Copiar o arquivo
    shutil.copy(origem, destino)

    print("Arquivo copiado com sucesso!")

    # Caminho do arquivo de entrada
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\ZBRINB006 Material.xlsx"

    # Colunas a serem mantidas
    colunas_desejadas = [
        "Nº NF-e",
        "Chave de acesso",
        "Data processamento",
        "CNPJ do emissor",
        "CNPJ destinatário",
        "Nome emissor",
        "Val.total c/impostos",
        "Data Vencimento",
        "Status global",
    ]

    # Ler o arquivo Excel
    df = pd.read_excel(caminho_arquivo)

    # Filtrar apenas as colunas desejadas
    df_filtrado = df[colunas_desejadas]

    # Salvar o arquivo com as colunas filtradas
    df_filtrado.to_excel(caminho_arquivo, index=False)

    print("Arquivo atualizado com sucesso!")

    # Caminhos dos arquivos
    pasta_erro = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - ERROS"
    arquivo_principal = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\ZBRINB006 Material.xlsx"

    # Leitura do arquivo principal
    df_principal = pd.read_excel(arquivo_principal, sheet_name="Sheet1")

    # Lista para armazenar os dataframes dos arquivos de erro
    dataframes_erro = []

    # Verifica e lê todos os arquivos na pasta de erros sem modificar os arquivos
    arquivos_na_pasta = os.listdir(pasta_erro)
    print(f"Arquivos encontrados na pasta: {arquivos_na_pasta}")

    for arquivo in arquivos_na_pasta:
        # Verifica se a extensão do arquivo é .xlsx
        if arquivo.lower().endswith(".xlsx"):
            caminho_arquivo = os.path.join(pasta_erro, arquivo)
            try:
                df_erro = pd.read_excel(caminho_arquivo)
                # Filtra colunas totalmente vazias ou preenchidas com NA
                df_erro = df_erro.dropna(axis=1, how="all")
                # Adiciona o dataframe à lista se não estiver vazio
                if not df_erro.empty:
                    dataframes_erro.append(df_erro)
                    print(f"Arquivo lido com sucesso: {arquivo}")
                else:
                    print(f"Arquivo {arquivo} está vazio após filtrar colunas.")
            except Exception as e:
                print(f"Erro ao ler o arquivo {arquivo}: {e}")
        else:
            print(f"Arquivo ignorado (não é .xlsx): {arquivo}")

    # Verifica se a lista de dataframes não está vazia
    if dataframes_erro:
        # Concatena todos os dataframes da lista em um único dataframe
        df_erro_combinado = pd.concat(dataframes_erro, ignore_index=True)

        # Mescla os dataframes com base na coluna de referência
        df_resultado = pd.merge(
            df_principal,
            df_erro_combinado[["Chave de acesso NF-e", "Local de negócios"]],
            left_on="Chave de acesso",
            right_on="Chave de acesso NF-e",
            how="left",
        )

        # Remove a coluna auxiliar após a mesclagem
        df_resultado = df_resultado.drop(columns=["Chave de acesso NF-e"])

        # Salva o dataframe atualizado no arquivo principal
        df_resultado.to_excel(arquivo_principal, sheet_name="Sheet1", index=False)
        print("Arquivo atualizado com sucesso.")
    else:
        print("Nenhum arquivo de erro foi encontrado ou lido.")

    # Caminhos dos arquivos
    file_celula_fiscal = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    file_zbrinb006 = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\ZBRINB006 Material.XLSX"

    # Carregar os dados apenas da aba "Célula Fiscal"
    df_celula_fiscal = pd.read_excel(file_celula_fiscal, sheet_name="Célula Fiscal")
    df_zbrinb006 = pd.read_excel(file_zbrinb006)

    # Garantir que as chaves de acesso são strings e estão sem espaços
    df_celula_fiscal["Chave de acesso"] = (
        df_celula_fiscal["Chave de acesso"].astype(str).str.strip()
    )
    df_zbrinb006["Chave de acesso"] = (
        df_zbrinb006["Chave de acesso"].astype(str).str.strip()
    )

    # Filtrar dados que ainda não estão no arquivo de destino com base na 'Chave de acesso'
    df_new_entries = df_zbrinb006[
        ~df_zbrinb006["Chave de acesso"].isin(df_celula_fiscal["Chave de acesso"])
    ]

    # Remover linhas completamente vazias
    df_new_entries = df_new_entries.dropna(how="all")

    # Verificar se há novos dados
    if not df_new_entries.empty:
        # Adicionar a coluna 'Data Atualização' com a data atual
        df_new_entries["Data Atualização"] = datetime.today().strftime("%d/%m/%Y")

        # Carregar o workbook existente
        wb = load_workbook(file_celula_fiscal)
        ws = wb["Célula Fiscal"]

        # Mapear a ordem das colunas de df_celula_fiscal na planilha
        header = [cell.value for cell in ws[1]]
        col_index = {header[i]: i + 1 for i in range(len(header))}

        # Adicionar a coluna 'Data Atualização' se não estiver presente
        if "Data Atualização" not in header:
            ws.cell(row=1, column=len(header) + 1, value="Data Atualização")
            col_index["Data Atualização"] = len(header) + 1

        # Encontre a primeira linha vazia após as linhas com fórmulas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            if all(cell.value is None for cell in row):
                start_row = row[0].row
                break
        else:
            start_row = ws.max_row + 1

        # Adicionar os novos dados à planilha existente
        for _, row in df_new_entries.iterrows():
            new_row = [None] * len(col_index)
            for col in df_new_entries.columns:
                if col in col_index:
                    new_row[col_index[col] - 1] = row[col]
            if "Data Atualização" in col_index:
                new_row[col_index["Data Atualização"] - 1] = row["Data Atualização"]
            for i, value in enumerate(new_row):
                ws.cell(row=start_row, column=i + 1, value=value)
            start_row += 1

        # Salvar as mudanças no arquivo
        wb.save(file_celula_fiscal)
        print("Dados adicionados com sucesso.")
    else:
        print("Nenhum novo dado para adicionar.")

    caminho_celula_fiscal = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    caminho_zbrinb006 = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\ZBRINB006 Material.XLSX"

    # Ler os arquivos Excel
    df_celula_fiscal = pd.read_excel(caminho_celula_fiscal, sheet_name="Célula Fiscal")
    df_zbrinb006 = pd.read_excel(
        caminho_zbrinb006, sheet_name="Sheet1"
    )  # Troque 'Sheet1' se a aba tiver outro nome

    # Verificar se as colunas existem
    if (
        "Chave de acesso" not in df_celula_fiscal.columns
        or "Chave de acesso" not in df_zbrinb006.columns
    ):
        raise ValueError("Coluna 'Chave de acesso' não encontrada em um dos arquivos.")

    # Encontrar chaves que não estão no arquivo ZBRINB006
    chaves_celula_fiscal = df_celula_fiscal["Chave de acesso"]
    chaves_zbrinb006 = df_zbrinb006["Chave de acesso"]

    # Preencher a coluna 'Concluído?' para as chaves não encontradas com "Concluído"
    df_celula_fiscal["Concluído?"] = df_celula_fiscal["Chave de acesso"].apply(
        lambda x: "Concluído" if x not in chaves_zbrinb006.values else ""
    )

    # Carregar o workbook existente e a planilha específica
    wb = load_workbook(caminho_celula_fiscal)
    ws = wb["Célula Fiscal"]

    # Mapear as colunas do DataFrame com as da planilha
    col_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    col_concluido_index = col_map.get("Concluído?")

    # Adicionar os dados atualizados ao Excel
    for idx, row in df_celula_fiscal.iterrows():
        ws.cell(row=idx + 2, column=col_concluido_index, value=row["Concluído?"])

    # Salvar as alterações
    wb.save(caminho_celula_fiscal)
    print("Processo concluído com sucesso!")

    formulas_cf()

    caminho_arquivo = r"G:\\Drives compartilhados\\Fiscal_Arquivo_de_Notas\\FISCAL_MATERIAL\\CONTROLE GRC - MATERIAIS\\Célula Fiscal GRC - Materiais.XLSX"

    # Carregar o workbook e a planilha
    wb = load_workbook(caminho_arquivo, data_only=False)
    ws = wb["Célula Fiscal"]

    # Encontrar o índice da coluna "Concluído?"
    header = [cell.value for cell in ws[1]]
    col_concluido_idx = header.index("Concluído?") + 1  # Índice da coluna "Concluído?"

    # Identificar e remover linhas onde o valor da coluna "Concluído?" é "Concluído"
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2, max_col=col_concluido_idx, values_only=False):
        if row[col_concluido_idx - 1].value == "Concluído":
            rows_to_delete.append(row[0].row)

    # Excluir linhas em ordem inversa para evitar problemas de reindexação
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    # Salvar as mudanças
    wb.save(caminho_arquivo)

    print("Linhas excluídas com sucesso.")

    formulas_cf()

    # Defina os caminhos das pastas
    folder_path_obras = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Obras"
    folder_path_suprimentos = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Suprimentos"

    # Exclua os arquivos na pasta 'Obras'
    files_obras = glob.glob(os.path.join(folder_path_obras, "*"))
    for file in files_obras:
        try:
            os.remove(file)
            print(f"Successfully deleted: {file}")
        except Exception as e:
            print(f"Error deleting {file}: {e}")

    # Exclua os arquivos na pasta 'Suprimentos'
    files_suprimentos = glob.glob(os.path.join(folder_path_suprimentos, "*"))
    for file in files_suprimentos:
        try:
            os.remove(file)
            print(f"Successfully deleted: {file}")
        except Exception as e:
            print(f"Error deleting {file}: {e}")

    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    caminho_temp = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\Célula Fiscal GRC - Materiais Temp.xlsx"
    criar_copia_temporaria(caminho_arquivo, caminho_temp)

    pasta_saida_obras = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Obras"
    pasta_saida_suprimentos = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Suprimentos"

    # Carregar a planilha temporária sem fórmulas
    try:
        df = pd.read_excel(caminho_temp, sheet_name="Célula Fiscal")
        print("Planilha carregada com sucesso.")
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        exit()

    # Garantir que as pastas de saída existam
    os.makedirs(pasta_saida_obras, exist_ok=True)
    os.makedirs(pasta_saida_suprimentos, exist_ok=True)

    # Filtrar e salvar os itens de "Obra"
    df_obra = df[df["Responsável - Etapa Atual"] == "Obra"]
    print(f"Total de itens de 'Obra' encontrados: {len(df_obra)}")
    if df_obra.empty:
        print("Nenhum item encontrado para 'Obra'.")
    else:
        for local in df_obra["Local de negócios"].dropna().unique():
            print(f"Processando local: {local}")
            df_filtro = df_obra[df_obra["Local de negócios"] == local]
            print(f"Total de itens para {local}: {len(df_filtro)}")
            nome_arquivo = (
                str(local).replace("/", "-").replace("\\", "-")
            )  # Substituir barras para evitar erros
            caminho_saida = os.path.join(pasta_saida_obras, f"{nome_arquivo}.xlsx")
            try:
                df_filtro.to_excel(caminho_saida, index=False)
                print(f"Arquivo salvo para {local} em {caminho_saida}")
            except Exception as e:
                print(f"Erro ao salvar o arquivo para {local}: {e}")
                continue

            # Verificar se o arquivo foi realmente salvo
            if os.path.exists(caminho_saida):
                print(f"Arquivo {caminho_saida} existe.")
            else:
                print(
                    f"Arquivo {caminho_saida} não foi encontrado após tentativa de salvamento."
                )
                continue

            # Ajustar layout do arquivo criado
            try:
                workbook = load_workbook(caminho_saida)
                sheet = workbook.active

                # Ajuste automático da largura das colunas
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Congelar a primeira linha
                sheet.freeze_panes = "A2"

                workbook.save(caminho_saida)
                print(f"Layout ajustado para {local}")
            except Exception as e:
                print(f"Erro ao ajustar o layout para {local}: {e}")

    # Filtrar e salvar os itens de "COMPRADOR" e "ASS TÉCNICA"
    df_suprimentos = df[df["Status"].isin(["COMPRADOR", "ASS TÉCNICA"])]
    if df_suprimentos.empty:
        print("Nenhum item encontrado para 'COMPRADOR' ou 'ASS TÉCNICA'.")
    else:
        caminho_saida_suprimentos = os.path.join(
            pasta_saida_suprimentos, "Suprimentos.xlsx"
        )
        df_suprimentos.to_excel(caminho_saida_suprimentos, index=False)
        print(f"Arquivo de suprimentos salvo em {caminho_saida_suprimentos}")

        # Ajustar layout do arquivo de "Suprimentos"
        try:
            workbook = load_workbook(caminho_saida_suprimentos)
            sheet = workbook.active

            # Ajuste automático da largura das colunas
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Congelar a primeira linha e a primeira coluna
            sheet.freeze_panes = "B2"

            # Ajustar a largura da coluna "Observação"
            observacao_column_letter = sheet["A"][0].column_letter
            sheet.column_dimensions[observacao_column_letter].width = 20

            # Ordenar as linhas pela coluna "Responsável - Etapa Atual"
            df_suprimentos_sorted = df_suprimentos.sort_values(
                by="Responsável - Etapa Atual"
            )
            df_suprimentos_sorted.to_excel(caminho_saida_suprimentos, index=False)

            workbook.save(caminho_saida_suprimentos)
            print("Layout ajustado para suprimentos")
        except Exception as e:
            print(f"Erro ao ajustar o layout para suprimentos: {e}")

    # Remover o arquivo temporário
    try:
        os.remove(caminho_temp)
        print("Arquivo temporário removido com sucesso.")
    except Exception as e:
        print(f"Erro ao remover o arquivo temporário: {e}")

    print("Processo concluído com sucesso.")

    # Caminho das pastas de saída para obras e suprimentos
    pasta_saida_obras = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Obras"
    pasta_saida_suprimentos = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Suprimentos"

    # Colunas para manter nos arquivos de obras
    colunas_obras = [
        "Status",
        "Observação",
        "Nº NF-e",
        "Chave de acesso",
        "Data processamento",
        "CNPJ do emissor",
        "CNPJ destinatário",
        "Nome emissor",
        "Val.total c/impostos",
        "Data Vencimento",
        "Status global",
        "Local de negócios",
    ]

    # Colunas para manter nos arquivos de suprimentos
    colunas_suprimentos = [
        "Responsável - Etapa Atual",
        "Data Vencimento",
        "Observação",
        "Local",
        "Nº NF-e",
        "Chave de acesso",
        "Data processamento",
        "CNPJ do emissor",
        "CNPJ destinatário",
        "Nome emissor",
        "Val.total c/impostos",
        "Status global",
        "Local de negócios",
    ]

    def ajustar_layout(caminho_arquivo):
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active

        # Ajuste automático da largura das colunas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            if column[0].value == "Observação":
                adjusted_width = 20
            else:
                adjusted_width = max_length + 2

            sheet.column_dimensions[column_letter].width = adjusted_width

        # Congelar a primeira linha e primeira coluna
        sheet.freeze_panes = sheet["B2"]
        workbook.save(caminho_arquivo)

    # Filtrar e salvar os arquivos de obras com as colunas desejadas
    for root, dirs, files in os.walk(pasta_saida_obras):
        for file in files:
            if file.endswith(".xlsx"):
                caminho_arquivo = os.path.join(root, file)
                df = pd.read_excel(caminho_arquivo)
                df_filtrado = df[colunas_obras]
                df_filtrado.to_excel(caminho_arquivo, index=False)
                ajustar_layout(caminho_arquivo)

    # Filtrar, ordenar e salvar o arquivo de suprimentos com as colunas desejadas
    for root, dirs, files in os.walk(pasta_saida_suprimentos):
        for file in files:
            if file.endswith(".xlsx"):
                caminho_arquivo = os.path.join(root, file)
                df = pd.read_excel(caminho_arquivo)
                df_filtrado = df[colunas_suprimentos]

                # Ordenar pelo valor na coluna 'Responsável - Etapa Atual'
                df_filtrado = df_filtrado.sort_values(by="Responsável - Etapa Atual")

                df_filtrado.to_excel(caminho_arquivo, index=False)
                ajustar_layout(caminho_arquivo)

    print("Processo de filtragem, ordenação e formatação concluído com sucesso.")

    # Filtrar, ordenar e salvar o arquivo de suprimentos com as colunas desejadas
    for root, dirs, files in os.walk(pasta_saida_suprimentos):
        for file in files:
            if file.endswith(".xlsx"):
                caminho_arquivo = os.path.join(root, file)
                df = pd.read_excel(caminho_arquivo)
                df_filtrado = df[colunas_suprimentos]

                # Ordenar pelo valor na coluna 'Data Vencimento' em ordem decrescente
                df_filtrado = df_filtrado.sort_values(
                    by="Data Vencimento", ascending=True
                )

                df_filtrado.to_excel(caminho_arquivo, index=False)
                ajustar_layout(caminho_arquivo)

    # Caminho do arquivo
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Suprimentos\Suprimentos.xlsx"

    # Carregar o workbook e a planilha
    wb = load_workbook(caminho_arquivo)
    ws = wb.active  # Supondo que a planilha de interesse é a ativa

    # Inserir a nova coluna na terceira posição (coluna C)
    ws.insert_cols(3)

    # Nomear a nova coluna
    cell = ws.cell(row=1, column=3)
    cell.value = "Justificativa Comprador"

    # Aplicar estilo de negrito
    bold_font = Font(bold=True)
    cell.font = bold_font

    # Aplicar bordas
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.border = thin_border

    # Salvar as mudanças
    wb.save(caminho_arquivo)


def email_suprimentos():
    usuario_gmail, senha_gmail = credenciais_login_email()

    # Verifica se as credenciais do Gmail são válidas
    if verificar_credenciais(usuario_gmail, senha_gmail):
        print("Sucesso: Credenciais válidas! Executando rotina...")
    else:
        messagebox.showerror(
            "Erro", "Credenciais inválidas. Verifique o e-mail e a senha."
        )
        close_process("saplogon.exe")
        return

    # Caminho do arquivo e nome da aba
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    nome_aba = "Dados"

    # Carregar a aba "Dados" do arquivo Excel
    df = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

    # Normalizar os nomes das colunas, removendo espaços e ajustando capitalização
    df.columns = df.columns.str.strip().str.lower()

    # Verificar se a coluna 'área' está presente
    if "área" in df.columns:
        # Filtrar os e-mails para as áreas específicas
        areas_desejadas = ["Suprimentos", "Ass Técnica", "Matriz"]
        df_filtrado = df[df["área"].isin(areas_desejadas)]

        # Obter os e-mails únicos e não nulos
        emails_filtrados = df_filtrado["e-mail"].dropna().unique()

        # Criar um DataFrame com os e-mails filtrados
        df_emails_filtrados = pd.DataFrame(emails_filtrados, columns=["E-mail"])

        # Configurar informações de e-mail
        email_remessa = usuario_gmail  # Usar o e-mail do usuário
        senha = senha_gmail  # Usar a senha do usuário
        arquivo_anexo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Suprimentos\Suprimentos.xlsx"

        # Criar a lista de e-mails
        emails_list = df_emails_filtrados["E-mail"].tolist()

        # Construir o e-mail
        msg = MIMEMultipart()
        msg["Subject"] = "Pedidos de compra PENDENTES de Ajuste"
        msg["From"] = email_remessa
        msg["To"] = ", ".join(emails_list)  # Enviar para todos os e-mails listados

        # Corpo do e-mail
        corpo_email = """
        <p>Prezados!<p>
        <p>Estamos enviando em anexo a lista de ajustes que são necessários para realizarmos o lançamento<p>
        <p>das NFs, verifiquem os itens do arquivo anexo para que possamos evitar atrasos nos pagamentos<p>
        <p>dessas NF-e, tendo em vista que isso tem impacto direto no nosso resultado!<p> 
        
        <p>OS AJUSTES REALIZADOS DEVEM SER AVISADOS PELO WORKPLACE!<p> 
        
        <p>Essa é uma mensagem automática, favor não responder!<p>

        <p>Link para acesso ao painel de Gestão de NFs:https://rogga.powerembedded.com.br/Organization/0a7254dc-2617-46ff-85b9-4d7b3c446bb4/Report/891e82f7-e9a7-4f3c-85c8-5884390334a7<p>
    
        <p>Link para download das NFs:https://drive.google.com/drive/folders/1EsjLOfnKe70qJMBdiH6mPPh8pMN60I8L?usp=drive_link<p>
        
        <p>Dúvidas, deverão entrar em contato diretamente com o fiscal material seguindo a divisão de obras.<p>
        <p>e-mail: fiscal.material@rogga.com.br<p>
        """
        msg.attach(MIMEText(corpo_email, "html"))

        # Adicionar o anexo
        with open(arquivo_anexo, "rb") as anexo:
            parte_anexo = MIMEBase("application", "octet-stream")
            parte_anexo.set_payload(anexo.read())
            encoders.encode_base64(parte_anexo)
            parte_anexo.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(arquivo_anexo)}",
            )
            msg.attach(parte_anexo)

        # Enviar o e-mail
        try:
            with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
                servidor.starttls()
                servidor.login(email_remessa, senha)
                servidor.sendmail(
                    email_remessa, emails_list, msg.as_string().encode("utf-8")
                )
            print("E-mail enviado para todos os destinatários.")
        except Exception as e:
            print(f"Falha ao enviar o e-mail: {e}")
    else:
        print("A coluna 'Área' não foi encontrada. Verifique o nome da coluna.")


def email_obras():
    usuario_gmail, senha_gmail = credenciais_login_email()

    # Verifica se as credenciais do Gmail são válidas
    if verificar_credenciais(usuario_gmail, senha_gmail):
        print("Sucesso: Credenciais válidas! Executando rotina...")
    else:
        messagebox.showerror(
            "Erro", "Credenciais inválidas. Verifique o e-mail e a senha."
        )
        close_process("saplogon.exe")
        return

    # Caminho do arquivo Excel
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"

    # Ler a aba "Dados" do arquivo Excel
    df = pd.read_excel(file_path, sheet_name="Dados")

    # Selecionar as colunas "E-mail" e "Local", filtrar as linhas onde essas colunas não estão vazias e onde "Local" não é "Matriz"
    df_emails_obras = df[["E-mail", "Local"]].dropna(subset=["E-mail", "Local"])
    df_emails_obras = df_emails_obras[df_emails_obras["Local"] != "Matriz"]

    # Converter a coluna "Local" para inteiros (remover casas decimais)
    df_emails_obras["Local"] = df_emails_obras["Local"].astype(int)

    # Lista de e-mails para adicionar em cópia (CC)
    cc_emails = [
        "andressa.fagundes@rogga.com.br",
        "nicole.soligo@rogga.com.br",
        "anderson.alexi@rogga.com.br",
        "marcelo.laude@rogga.com.br",
    ]

    # Caminho da pasta onde os arquivos estão localizados
    pasta_obras = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Envios\Obras"

    # Iterar sobre as linhas do DataFrame df_emails_obras
    for index, row in df_emails_obras.iterrows():
        # Construir o nome base do arquivo baseado na coluna "Local"
        nome_arquivo_base = str(
            int(row["Local"])
        )  # Garantir que seja string de inteiro
        # Listar todos os arquivos na pasta
        arquivos_na_pasta = os.listdir(pasta_obras)
        # Procurar por um arquivo que comece com nome_arquivo_base e termine com ".xlsx"
        arquivos_encontrados = [
            arquivo
            for arquivo in arquivos_na_pasta
            if arquivo.startswith(nome_arquivo_base) and arquivo.endswith(".xlsx")
        ]

        if arquivos_encontrados:
            # Considerar o primeiro arquivo encontrado
            caminho_arquivo = os.path.join(pasta_obras, arquivos_encontrados[0])
            # Resto do código para envio de e-mail...
            # Separar os e-mails por vírgula e remover espaços extras
            to_email_list = [email.strip() for email in row["E-mail"].split(",")]
            # Configurar informações de e-mail
            email_remessa = usuario_gmail  # Usar o e-mail do usuário
            senha = senha_gmail  # Usar a senha do usuário
            # Construir um novo objeto MIMEMultipart para cada e-mail
            msg = MIMEMultipart()
            msg["Subject"] = (
                "Notas fiscais de materiais disponíveis para lançamento, Sem Saldo e Sem Pedido"
            )
            msg["From"] = email_remessa
            msg["To"] = ", ".join(
                to_email_list
            )  # Agora lista todos os destinatários corretamente
            msg["CC"] = ", ".join(cc_emails)

            # Corpo do e-mail
            corpo_email = """
            <p>Prezados!<p>
            <p>Estamos enviando em anexo a lista de notas fiscais de material que estão disponíveis para lançamento via GRC, MIGO e NFs sem saldo. Verifiquem os itens do arquivo anexo,<p>
            <p>para que possamos evitar atrasos nos pagamentos dessas NF-e, tendo em vista que isso tem impacto direto no nosso resultado!<p>
            <p>Obs.: Lançar apenas as notas que já chegaram em obra ou matriz<p> 
            <p>Monitor portaria: https://portais.rogga.com.br:50000/sap/bc/webdynpro/xnfe/gatekeeper_workplace#<p> 
            <p>Monitor logístico: https://portais.rogga.com.br:50000/sap/bc/webdynpro/xnfe/nfe_log_workplace#<p>
            <p>Link para acesso ao painel de Gestão de NFs:https://rogga.powerembedded.com.br/Organization/0a7254dc-2617-46ff-85b9-4d7b3c446bb4/Report/d1913438-8abf-4155-8642-c2d1308be92c<p>
            <p>Essa é uma mensagem automática, favor não responder!<p>
            <p>Dúvidas entrar em contato no e-mail: fiscal.material@rogga.com.br<p>
            """
            msg.attach(MIMEText(corpo_email, "html"))

            # Anexar o arquivo ao e-mail
            with open(caminho_arquivo, "rb") as anexo:
                parte_anexo = MIMEBase("application", "octet-stream")
                parte_anexo.set_payload(anexo.read())
                encoders.encode_base64(parte_anexo)
                parte_anexo.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(caminho_arquivo)}",
                )
                msg.attach(parte_anexo)

            # Combinar lista de destinatários e cópias para envio
            all_recipients = to_email_list + cc_emails

            # Enviar o e-mail
            try:
                with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
                    servidor.starttls()
                    servidor.login(email_remessa, senha)
                    servidor.sendmail(
                        email_remessa, all_recipients, msg.as_string().encode("utf-8")
                    )
                print(
                    f'E-mail para {row["E-mail"]} enviado com o arquivo {arquivos_encontrados[0]}'
                )
            except Exception as e:
                print(f"Falha ao enviar o e-mail para {row['E-mail']}: {e}")
        else:
            print(f"O arquivo listado não existe para o Local {nome_arquivo_base}")


def SAP_Extra():
    # Definindo os caminhos dos arquivos
    origem = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal GRC - Materiais.XLSX"
    destino = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Célula Fiscal\Célula Fiscal GRC - Materiais.XLSX"

    # Verifica se o arquivo de destino já existe
    if os.path.exists(destino):
        os.remove(destino)  # Remove o arquivo existente

    # Copia o arquivo de origem para o destino
    shutil.copy2(origem, destino)

    print(f"Arquivo copiado com sucesso para {destino}")

    # Caminho da pasta
    pasta = sap_gui_path

    # Listar todos os arquivos na pasta
    arquivos = os.listdir(pasta)

    # Apagar cada arquivo
    for arquivo in arquivos:
        caminho_arquivo = os.path.join(pasta, arquivo)
        if os.path.isfile(caminho_arquivo):  # Verifica se é um arquivo
            os.remove(caminho_arquivo)

    # Fazer login no SAP
    fazer_login()
    # INFORMAÇÕES SAP
    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "zfi017"
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]/usr/ctxtP_VARIA").text = "/LT_ZFI017"
        session.findById("wnd[0]/usr/ctxtS_BUPLA-LOW").text = "0001"
        session.findById("wnd[0]/usr/ctxtS_BUPLA-HIGH").text = "0072"
        session.findById("wnd[0]/usr/txtP_GJAHR").text = ano_corrente
        session.findById("wnd[0]/usr/txtS_MONAT-LOW").text = "1"
        session.findById("wnd[0]/usr/txtS_MONAT-HIGH").text = "12"

        # Obter a data de hoje
        hoje = datetime.now()

        # Verificar se hoje é o primeiro dia do mês
        if hoje.day == 1:
            # Caso seja dia 1, D1 será o primeiro dia do mês anterior
            primeiro_dia_mes_atual = hoje.replace(day=1)
            ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
            D1 = ultimo_dia_mes_anterior.replace(day=1).strftime("%d%m%Y")

            # DF será o último dia do mês anterior
            DF = ultimo_dia_mes_anterior.strftime("%d%m%Y")

            # DF_F será o mês e ano do mês anterior
            DF_F = ultimo_dia_mes_anterior.strftime("%m.%Y")
        else:
            # Caso contrário, D1 será o primeiro dia do mês atual
            D1 = hoje.replace(day=1).strftime("%d%m%Y")

            # DF será o dia de hoje
            DF = hoje.strftime("%d%m%Y")

            # DF_F será o mês e ano do mês atual
            DF_F = hoje.strftime("%m.%Y")

        session.findById("wnd[0]/usr/ctxtS_BUDAT-LOW").text = D1
        session.findById("wnd[0]/usr/ctxtS_BUDAT-HIGH").text = DF
        session.findById("wnd[0]/usr/ctxtS_BUDAT-HIGH").setFocus
        session.findById("wnd[0]/usr/ctxtS_BUDAT-HIGH").caretPosition = 6
        session.findById("wnd[0]").sendVKey(8)
        session.findById(
            "wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell"
        ).currentCellColumn = "STCD1"
        session.findById(
            "wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell"
        ).selectedRows = "0"
        session.findById(
            "wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell"
        ).contextMenu()
        session.findById(
            "wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[1]/shell"
        ).selectContextMenuItem("&XXL")
        session.findById("wnd[1]").sendVKey(0)
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = sap_gui_path
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZFI017_" + DF_F + ".XLSX"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 6
        session.findById("wnd[1]").sendVKey(11)
        session.findById("wnd[0]").sendVKey(3)

        close_process("saplogon.exe")

        pasta_origem_input = sap_gui_path
        pasta_destino_input = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZFI017"

        processar_arquivos(pasta_origem_input, pasta_destino_input)

    except:
        close_process("saplogon.exe")


def copiar_e_congelar_arquivo():
    # Caminho do arquivo de origem e destino
    origem = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\GRC\ZBRINB006 - GERAL\ZBRINB006 Serviço.XLSX"
    destino = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_MATERIAL\CONTROLE GRC - MATERIAIS\Notas de Serviço.XLSX"

    # Copia o arquivo
    shutil.copyfile(origem, destino)

    # Carrega o arquivo copiado para aplicar o congelamento
    wb = load_workbook(destino)
    ws = wb.active

    # Congela a célula D2
    ws.freeze_panes = "D2"

    # Salva as alterações
    wb.save(destino)
    wb.close()


def rotina1():
    GRC_ETL()
    salvar_copia_celula_fiscal()
    compilar_historico()
    SAP_Extra()
    copiar_e_congelar_arquivo()
    time.sleep(20)
    nf_recebida()


def rotina2():
    email_suprimentos()


def rotina3():
    email_obras()


def abrir_link():
    url = "https://docs.google.com/spreadsheets/d/1JPhouDDSGI3fBqSlICmVl7IHG4KhN1UG/edit?usp=drive_link&ouid=103952633831854335180&rtpof=true&sd=true"
    webbrowser.open(url)


def abrir_senha():
    url = "https://myaccount.google.com/apppasswords?pli=1&rapt=AEjHL4PAtIkOxG1y_l4Zs_tgi0lh7PJbJStXnwiDi9wWv4sfSY0Js-ybon-yqXilW9AW62nKm9owXINq_K4VEhHtNoC7YQstlwRCx5o8_k3rTx_jr-Zcp4I"
    webbrowser.open(url)


# Criando a janela principal
root = tk.Tk()
root.title("GRC")
root.geometry("235x450")  # Ajuste para comportar melhor os widgets
root.configure(bg="#f2f2f2")
# root.resizable(False, False)  # Impede o redimensionamento da janela

# Fonte personalizada
root.option_add("*Font", "Amiko 10")

# Título centralizado com quebra de linha
label_titulo = tk.Label(
    root,
    text="Rotinas GRC",
    font=("Amiko", 10, "bold"),
    bg="#f2f2f2",
    fg="#b23a48",
    wraplength=240,  # Ajuste a largura conforme necessário
)
label_titulo.pack(pady=(10, 10))

# Frame principal para centralizar os elementos
frame = tk.Frame(root, bg="#f2f2f2")
frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=5)

# Configuração de estilo para os widgets
input_style = {
    "font": ("Amiko", 10),
    "bg": "#ffffff",
    "bd": 0,
    "highlightthickness": 1,
    "highlightbackground": "#d1d1d1",
    "highlightcolor": "#b23a48",
    "relief": "flat",
}

label_style = {"bg": "#f2f2f2", "font": ("Amiko", 10)}
button_style = {
    "font": ("Amiko", 10),
    "bg": "#b23a48",
    "fg": "white",
    "bd": 0,
    "activebackground": "#8c2e39",
    "relief": "flat",
    "cursor": "hand2",
}

# Criando rótulos e entradas
tk.Label(frame, text="Usuário (SAP)", **label_style).grid(
    row=0, column=0, sticky="w", pady=(2, 1)
)
entry_usuario_sap = tk.Entry(frame, **input_style, width=25)
entry_usuario_sap.grid(row=1, column=0, sticky="ew", pady=(0, 5))

tk.Label(frame, text="Senha (SAP)", **label_style).grid(
    row=2, column=0, sticky="w", pady=(2, 1)
)
entry_senha_sap = tk.Entry(frame, **input_style, width=25, show="*")
entry_senha_sap.grid(row=3, column=0, sticky="ew", pady=(0, 5))

# Centralizar o botão
button_executar = tk.Button(
    frame, text="Atualização", command=rotina1, **button_style, width=25
)
button_executar.grid(row=4, column=0, pady=(10, 10), sticky="ew")


tk.Label(frame, text="Usuário (Gmail)", **label_style).grid(
    row=5, column=0, sticky="w", pady=(2, 1)
)
entry_usuario_gmail = tk.Entry(frame, **input_style, width=25)
entry_usuario_gmail.grid(row=6, column=0, sticky="ew", pady=(0, 5))

tk.Label(frame, text="Senha (Gmail)", **label_style).grid(
    row=7, column=0, sticky="w", pady=(2, 1)
)
entry_senha_gmail = tk.Entry(frame, **input_style, width=25, show="*")
entry_senha_gmail.grid(row=8, column=0, sticky="ew", pady=(0, 10))


# Centralizar o botão
button_executar = tk.Button(
    frame, text="E-mail Suprimentos", command=rotina2, **button_style, width=25
)
button_executar.grid(row=9, column=0, pady=(10, 10), sticky="ew")

# Centralizar o botão
button_executar = tk.Button(
    frame, text="E-mail Obras", command=rotina3, **button_style, width=25
)
button_executar.grid(row=10, column=0, pady=(10, 10), sticky="ew")

# Rótulo para abrir o link da senha
label_link_senha = tk.Label(
    frame,
    text="Senha para acesso Gmail",
    fg="blue",
    cursor="hand2",
    bg="#f2f2f2",
    font=("Amiko", 10, "underline"),
)
label_link_senha.grid(row=11, column=0, pady=(0, 10), sticky="ew")
label_link_senha.bind("<Button-1>", lambda e: abrir_senha())

# Rótulo para abrir o link da planilha
label_link = tk.Label(
    frame,
    text="Célula Fiscal GRC",
    fg="blue",
    cursor="hand2",
    bg="#f2f2f2",
    font=("Amiko", 10, "underline"),
)
label_link.grid(row=12, column=0, pady=(0, 20), sticky="ew")
label_link.bind("<Button-1>", lambda e: abrir_link())


# Configurar o comportamento de hover para os botões
def on_enter(e):
    e.widget["bg"] = "#8c2e39"


def on_leave(e):
    e.widget["bg"] = "#b23a48"


button_executar.bind("<Enter>", on_enter)
button_executar.bind("<Leave>", on_leave)


# Inicializa a interface
root.mainloop()

# %%
