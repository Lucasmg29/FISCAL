import os
import shutil
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table
import win32com.client
import psutil
import csv
import glob
import subprocess
import time
import tkinter as tk
from tkinter import messagebox
import shlex


def organizar_arquivos(pasta_origem, pasta_destino):
    # Obtém a data atual
    hoje = datetime.now()
    nome_pasta_ano = hoje.strftime("%Y")  # Formato: AAAA
    nome_pasta_mes = hoje.strftime("%m.%Y")  # Formato: MM.AAAA

    # Cria a pasta do ano e do mês no destino
    caminho_pasta_ano = os.path.join(pasta_destino, nome_pasta_ano)
    caminho_pasta_mes = os.path.join(caminho_pasta_ano, nome_pasta_mes)

    # Cria as pastas, se não existirem
    os.makedirs(caminho_pasta_mes, exist_ok=True)

    # Move os arquivos da pasta de origem para a pasta do mês
    for arquivo in os.listdir(pasta_origem):
        caminho_arquivo = os.path.join(pasta_origem, arquivo)
        if os.path.isfile(caminho_arquivo):  # Verifica se é um arquivo
            shutil.move(caminho_arquivo, os.path.join(caminho_pasta_mes, arquivo))


def fazer_login():

    # CONECTAR SAP
    caminho_executavel_sap = r"C:\Program Files\SAP\FrontEnd\SAPGUI\saplogon.exe"
    if not os.path.exists(caminho_executavel_sap):
        caminho_executavel_sap = (
            r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        )

    print(f"Caminho do executável: {caminho_executavel_sap}")

    if not os.path.exists(caminho_executavel_sap):
        print("O arquivo não foi encontrado. Verifique o caminho.")
        messagebox.showerror(
            "Erro", "O executável do SAP GUI não foi encontrado. Verifique o caminho."
        )
        return

    # Use shlex.split para tratar corretamente o caminho com espaços
    comando = shlex.split(f'"{caminho_executavel_sap}"')
    print(f"Comando para execução: {comando}")

    try:
        subprocess.Popen(comando)
    except Exception as e:
        print(f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao iniciar o SAP GUI: {e}")
        return

    time.sleep(3)

    usuario = entry_usuario.get()
    senha = entry_senha.get()
    loc1 = entry_loc1.get()
    loc2 = entry_loc2.get()
    data1 = entry_data1.get()
    data2 = entry_data2.get()

    try:
        subprocess.Popen(comando)
    except Exception as e:
        print(f"Ocorreu um erro ao iniciar o SAP GUI: {e}")

    time.sleep(3)
    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    try:
        connection = application.OpenConnection("NOVO PVR - Produção SAP ECC", True)
    except Exception as e:
        print(
            f"Conexão 'PVR - Produção (Externo)' não encontrada. Tentando 'PVR - Produção (Interno)'..."
        )
        try:
            connection = application.OpenConnection("NOVO PVR - Produção SAP ECC", True)
        except Exception as e:
            print(f"Ocorreu um erro ao abrir a conexão: {e}")

    session = connection.Children(0)

    try:
        # Preencha as informações de cliente, usuário e senha
        session.findById("wnd[0]/usr/txtRSYST-MANDT").text = "300"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = usuario
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = senha
        session.findById("wnd[0]").sendVKey(0)
    except Exception as e:
        print(f"Ocorreu um erro durante a autenticação no SAP: {e}")

    # FUNÇÕES SAP

    sapguiauto = win32com.client.GetObject("SAPGUI")
    application = sapguiauto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "zfi017"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/ctxtS_BUPLA-LOW").text = loc1
    session.findById("wnd[0]/usr/ctxtS_BUPLA-HIGH").text = loc2
    session.findById("wnd[0]/usr/txtP_GJAHR").text = ano_atual
    session.findById("wnd[0]/usr/txtS_MONAT-LOW").text = mes_atual
    session.findById("wnd[0]/usr/txtS_MONAT-HIGH").text = mes_atual
    session.findById("wnd[0]/usr/ctxtS_BUDAT-LOW").text = data1
    session.findById("wnd[0]/usr/ctxtS_BUDAT-HIGH").text = data2
    session.findById("wnd[0]/usr/ctxtS_NFTYPE-LOW").text = "ZS"
    session.findById("wnd[0]/usr/ctxtS_NFTYPE-HIGH").text = "ZT"
    session.findById("wnd[0]/usr/ctxtP_VARIA").text = "/ZFI017_PY2"
    session.findById("wnd[0]/usr/ctxtP_VARIA").setFocus
    session.findById("wnd[0]/usr/ctxtP_VARIA").caretPosition = 10
    session.findById("wnd[0]").sendVKey(8)

    try:
        session.findById("wnd[0]").sendVKey(45)
        
    except:
        session.findById("wnd[1]").sendVKey(0)
        session.findById("wnd[1]").sendVKey(0)
        session.findById("wnd[1]").sendVKey(0)
        session.findById("wnd[0]").sendVKey(45)

    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = (
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência"
    )
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZFI017.txt"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 6
    session.findById("wnd[1]").sendVKey(11)

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nZSD001"
    session.findById("wnd[0]").sendVKey(0)

    session.findById("wnd[0]/usr/cntlCC_INI/shellcont/shell").pressToolbarButton(
        "&COL0"
    )
    session.findById(
        "wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell"
    ).currentCellRow = 1
    session.findById(
        "wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell"
    ).selectedRows = "1"
    session.findById(
        "wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/cntlCONTAINER1_LAYO/shellcont/shell"
    ).doubleClickCurrentCell()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[0]/usr/cntlCC_INI/shellcont/shell").firstVisibleColumn = (
        "OBRA"
    )

    session.findById("wnd[0]/usr/cntlCC_INI/shellcont/shell").pressToolbarContextButton(
        "&MB_EXPORT"
    )
    session.findById("wnd[0]/usr/cntlCC_INI/shellcont/shell").selectContextMenuItem(
        "&XXL"
    )
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = (
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência"
    )
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "MATRIZ.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 6
    session.findById("wnd[1]").sendVKey(11)

    data_atual_formatada = data1
    data_anterior_formatada = data2

    # Exibe os resultados
    print("Data atual (considerando o dia anterior):", data_atual_formatada)
    print("Data do dia anterior:", data_anterior_formatada)
    print("Mês atual:", mes_atual)
    print("Ano atual:", ano_atual)


def close_process(nome_processo):
    for proc in psutil.process_iter(["pid", "name"]):
        if nome_processo.lower() in proc.info["name"].lower():
            try:
                processo = psutil.Process(proc.info["pid"])
                processo.terminate()  # ou processo.kill() para forçar o fechamento
                print(f'{proc.info["name"]} (PID: {proc.info["pid"]}) foi fechado.')
            except psutil.NoSuchProcess:
                print(f'Erro: Processo {proc.info["name"]} não encontrado.')
            except psutil.AccessDenied:
                print(f'Erro: Acesso negado para fechar {proc.info["name"]}.')


data_atual = datetime.now() - timedelta(days=1)

if data_atual.weekday() == 6:  # 0 é segunda-feira, 6 é domingo
    # Ajusta a data atual para domingo
    data_atual = data_atual

    # Obtém a data da sexta-feira anterior
    data_anterior = data_atual - timedelta(days=2)
else:
    # Caso contrário, a data anterior é simplesmente um dia antes da data atual
    data_anterior = datetime.now() - timedelta(days=1)

# Obtém o valor do mês atual
mes_atual = data_atual.month

# Obtém o ano atual
ano_atual = data_atual.year


def move_coluna(nome_coluna, nova_posicao, caminho_arquivo):
    # Carregar o arquivo Excel
    df = pd.read_excel(caminho_arquivo)

    # Verificar se a coluna existe
    if nome_coluna not in df.columns:
        print(f"A coluna '{nome_coluna}' não existe no arquivo.")
        return

    # Obter o índice atual da coluna
    indice_coluna = df.columns.get_loc(nome_coluna)

    # Remover a coluna do DataFrame
    coluna_removida = df.pop(nome_coluna)

    # Inserir a coluna na nova posição
    df.insert(nova_posicao, nome_coluna, coluna_removida)

    # Salvar o DataFrame modificado de volta no arquivo Excel
    df.to_excel(caminho_arquivo, index=False)


def executar_rotina():
    close_process("saplogon.exe")
    fazer_login()
    # Fechar o SAP
    time.sleep(5)
    close_process("saplogon.exe")

    try:
        close_process("excel.exe")
    except:
        pass

    input_folder = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência"

    # Lista todos os arquivos TXT na pasta de entrada
    txt_files = glob.glob(os.path.join(input_folder, "*.txt"))

    for txt_file in txt_files:
        # Construa o caminho de saída alterando a extensão para .xlsx
        output_file = txt_file.replace(".txt", ".xlsx")

        # Remova o arquivo de saída se ele já existir
        if os.path.exists(output_file):
            os.remove(output_file)

        # Crie um novo arquivo XLSX
        wb = openpyxl.Workbook()
        ws = wb.active

        # Leitura do arquivo TXT e gravação no arquivo XLSX
        with open(txt_file, "r") as data:
            reader = csv.reader(data, delimiter="|")
            for row in reader:
                ws.append(row)

        # Salvar o arquivo XLSX
        wb.save(output_file)

        # Remover o arquivo TXT após a conversão
        os.remove(txt_file)

    diretorio = input_folder

    # Lista todos os arquivos Excel no diretório
    arquivos_excel = [
        arquivo for arquivo in os.listdir(diretorio) if arquivo.endswith(".xlsx")
    ]

    # Itera sobre cada arquivo Excel
    for arquivo in arquivos_excel:
        # Carrega o arquivo Excel em um DataFrame do pandas
        caminho_arquivo = os.path.join(diretorio, arquivo)
        df = pd.read_excel(caminho_arquivo)

        # Remove os espaços em branco em excesso de cada célula
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

        # Salva o DataFrame modificado de volta para o arquivo Excel
        df.to_excel(caminho_arquivo, index=False)

    # Diretório contendo os arquivos xlsx
    folder_path = input_folder

    # Lista de arquivos que requerem a exclusão de 6 linhas
    files_with_six_rows = []

    # Iterar sobre todos os arquivos na pasta
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):  # Verificar se é um arquivo Excel
            file_path = os.path.join(folder_path, filename)

            # Carregar o workbook
            workbook = load_workbook(file_path)

            # Selecionar a planilha ativa ou uma específica se necessário
            sheet = workbook.active

            # Excluir a primeira coluna
            sheet.delete_cols(1)

            # Verificar se é um dos arquivos que requerem a exclusão de 6 linhas
            if filename in files_with_six_rows:
                # Excluir as seis primeiras linhas
                for _ in range(6):
                    sheet.delete_rows(1)
            else:
                # Excluir as cinco primeiras linhas
                for _ in range(6):
                    sheet.delete_rows(1)

            # Salvar o arquivo
            workbook.save(file_path)

    # Caminho para o arquivo Excel
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carrega o arquivo Excel
    df = pd.read_excel(file_path)

    # Exclui as linhas onde a célula da segunda coluna (índice 1) está vazia
    df_cleaned = df.dropna(subset=[df.columns[1]])

    # Salva o dataframe limpo de volta no arquivo Excel
    df_cleaned.to_excel(file_path, index=False)

    # Carregar o arquivo Excel
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar o workbook
    workbook = load_workbook(filename=caminho_arquivo)

    # Selecionar a primeira planilha
    sheet = workbook.active

    # Alterar os nomes das células A1, B1 e C1
    sheet["F1"] = "Vl.Contábil"
    sheet["K1"] = "IR Retido"
    sheet["L1"] = "INSS Retido"
    sheet["M1"] = "ISS Retido"

    # Salvar as alterações de volta no arquivo Excel
    workbook.save(filename=caminho_arquivo)

    # Diretório contendo os arquivos xlsx
    folder_path = input_folder

    # Iterar sobre todos os arquivos na pasta
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):  # Verificar se é um arquivo Excel
            file_path = os.path.join(folder_path, filename)

            # Carregar o workbook
            workbook = openpyxl.load_workbook(file_path)

            # Iterar sobre as planilhas no workbook
            for sheet in workbook:
                # Ajustar tanto a coluna B quanto a coluna C
                for coluna_a_ajustar in ["F", "G", "K", "L", "M"]:
                    # Iterar sobre as células na coluna a ser ajustada
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet[coluna_a_ajustar + str(row)]
                        # Remover pontos como separadores de milhares
                        if isinstance(cell.value, str):
                            value = cell.value.replace(".", "")
                            # Se o último caractere for um hífen (-), movê-lo para antes da parte numérica
                            if value.endswith("-"):
                                value = "-" + value[:-1]
                            # Converter o valor da célula para float e aplicar o formato numérico
                            try:
                                value = float(
                                    value.replace(",", ".")
                                )  # Substituir a vírgula por ponto
                                cell.value = value
                                cell.number_format = "#,##0.00"
                            except ValueError:
                                # Se a conversão falhar, pular a célula
                                continue

            # Salvar o arquivo
            workbook.save(file_path)

    # Caminho do arquivo
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar a pasta de trabalho
    workbook = openpyxl.load_workbook(file_path)

    # Selecionar a aba ativa (ou você pode especificar uma aba pelo nome)
    sheet = workbook.active

    # Alterar o valor da célula 'E1'
    sheet["E1"] = "CPF/CNPJ"

    # Salvar o arquivo
    workbook.save(file_path)

    print("A célula 'E1' foi alterada para 'CPF/CNPJ'")

    # Carregar o arquivo F.19
    zfi017_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    zfi017 = pd.read_excel(zfi017_path)

    # Criar a tabela dinâmica
    pivot_table = pd.pivot_table(
        zfi017, index=["Nro NF", "Nome"], values=["Vl.Contábil"], aggfunc="sum"
    ).reset_index()

    # Carregar o arquivo Excel com openpyxl
    wb = load_workbook(zfi017_path)
    ws = wb.create_sheet(title="Tabela Dinâmica")

    # Adicionar os dados da tabela dinâmica ao Excel
    for r_idx, row in enumerate(
        dataframe_to_rows(pivot_table, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Adicionar as colunas na tabela dinâmica
    headers = [
        "Dt.lçto.",
        "Data doc.",
        "CPF/CNPJ",
        "Valor",
        "NCM",
        "Local.neg.",
        "Div",
        "IR Retido",
        "INSS Retido",
        "ISS Retido",
        "Nº doc.",
    ]
    for col_idx, header in enumerate(headers, len(pivot_table.columns) + 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Mapeamento de colunas do DataFrame para as colunas no Excel
    column_mapping = {
        "Dt.lçto.": 4,
        "Data doc.": 5,
        "CPF/CNPJ": 6,
        "Valor": 7,
        "NCM": 8,
        "Local.neg.": 9,
        "Div": 10,
        "IR Retido": 11,
        "INSS Retido": 12,
        "ISS Retido": 13,
        "Nº doc.": 14,
    }

    # Preencher os dados adicionais nas colunas correspondentes
    for r_idx, row in pivot_table.iterrows():
        matching_rows = zfi017[
            (zfi017["Nro NF"] == row["Nro NF"]) & (zfi017["Nome"] == row["Nome"])
        ]
        if not matching_rows.empty:
            for col_name, col_idx in column_mapping.items():
                ws.cell(
                    row=r_idx + 2, column=col_idx, value=matching_rows[col_name].iloc[0]
                )

    # Ordenar as linhas pela coluna "Local.neg."
    data_rows = list(
        ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
            values_only=True,
        )
    )
    sorted_rows = sorted(
        data_rows, key=lambda x: x[4]
    )  # Index 4 corresponde à coluna "Local.neg."

    # Reescrever as linhas ordenadas na planilha
    for row_idx, row in enumerate(sorted_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adicionar uma tabela ao Excel
    tab = Table(displayName="TabelaDinamica", ref=ws.dimensions)
    ws.add_table(tab)

    # Deletar a aba "Sheet1" se existir
    if "Sheet1" in wb.sheetnames:
        std = wb["Sheet1"]
        wb.remove(std)

    # Renomear a aba "Tabela Dinâmica" para "Sheet1"
    ws.title = "Sheet1"

    # Salvar as alterações no arquivo Excel
    wb.save(zfi017_path)

    # Exemplo de uso:
    nome_coluna = "Vl.Contábil"  # Nome da coluna que será movida
    nova_posicao = 5  # Nova posição da coluna
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"  # Caminho do arquivo Excel

    move_coluna(nome_coluna, nova_posicao, caminho_arquivo)

    # Exemplo de uso:
    nome_coluna = "Nome"  # Nome da coluna que será movida
    nova_posicao = 3  # Nova posição da coluna
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"  # Caminho do arquivo Excel

    move_coluna(nome_coluna, nova_posicao, caminho_arquivo)

    # Carregar o arquivo Excel
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar o workbook
    workbook = load_workbook(filename=caminho_arquivo)

    # Selecionar a primeira planilha
    sheet = workbook.active

    # Alterar os nomes das células A1, B1 e C1
    sheet["G1"] = "BC do INSS"

    # Salvar as alterações de volta no arquivo Excel
    workbook.save(filename=caminho_arquivo)

    # Carregar o arquivo Excel
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar o workbook
    workbook = load_workbook(filename=caminho_arquivo)

    # Selecionar a primeira planilha
    sheet = workbook.active

    # Alterar os nomes das células A1, B1 e C1
    sheet["O1"] = "ANEXADO"
    sheet["P1"] = "INSS PROVA REAL"
    sheet["Q1"] = "LOCAL DA PRESTAÇÃO"

    # Salvar as alterações de volta no arquivo Excel
    workbook.save(filename=caminho_arquivo)

    # Caminho do arquivo Excel
    matriz_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\MATRIZ.xlsx"

    # Carregar o arquivo Excel em um DataFrame
    df = pd.read_excel(matriz_path)

    # Excluir linhas com células vazias na primeira coluna
    df = df.dropna(subset=[df.columns[0]])

    # Remover duplicatas considerando as colunas 'Divisão' e 'Nome Empto.'
    df_cleaned = df.drop_duplicates(subset=["Divisão", "Nome Empto."])

    # Salvar o DataFrame limpo de volta ao arquivo Excel
    df_cleaned.to_excel(matriz_path, index=False)

    print(
        f"Remoção de duplicidades completa e linhas vazias na primeira coluna excluídas. O arquivo foi salvo em {matriz_path}."
    )

    # Carregar os arquivos
    df_zfi017 = pd.read_excel(
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    )
    df_matriz = pd.read_excel(
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\MATRIZ.xlsx"
    )

    # Mesclar os dataframes com base na coluna 'Div' em df_zfi017 e 'Divisão' em df_matriz
    df_merged = pd.merge(
        df_zfi017,
        df_matriz[["Divisão", "Nome Empto."]],
        left_on="Div",
        right_on="Divisão",
        how="left",
    )

    # Adicionar 'Fábrica' para os itens com '1016' na coluna 'Div' de ZFI017
    df_merged.loc[df_merged["Div"] == 1016, "Nome Empto."] = "Fábrica"

    # Adicionar 'Fábrica' para os itens com '1016' na coluna 'Div' de ZFI017
    df_merged.loc[df_merged["Div"] == 1001, "Nome Empto."] = "Matriz"

    # Excluir a coluna 'Divisão' que foi usada apenas para mesclar os dataframes
    df_merged.drop(columns="Divisão", inplace=True)

    # Salvar o resultado de volta no arquivo ZFI017.xlsx
    df_merged.to_excel(
        r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx",
        index=False,
    )

    # Caminho do arquivo Excel
    matriz_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar o arquivo Excel em um DataFrame
    df = pd.read_excel(matriz_path)

    # Remover duplicatas considerando as colunas 'Divisão' e 'Nome Empto.'
    df_cleaned = df.drop_duplicates(subset=["Nro NF", "Nome"])

    # Salvar o DataFrame limpo de volta ao arquivo Excel
    df_cleaned.to_excel(matriz_path, index=False)

    print(
        f"Remoção de duplicidades completa e linhas vazias na primeira coluna excluídas. O arquivo foi salvo em {matriz_path}."
    )

    # Ler o arquivo Excel, especificando que os nomes das colunas estão na segunda linha
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    dados_excel = pd.read_excel(caminho_arquivo, header=0)

    # Ordenar o DataFrame pela coluna 'Local.neg.'
    dados_excel = dados_excel.sort_values(by="Local.neg.")

    # Iterar sobre os valores únicos da coluna "Nome Empto." e "Local.neg."
    for nome_empto, grupo in dados_excel.groupby(["Nome Empto.", "Local.neg."]):
        # Criar um novo DataFrame para cada grupo
        df_grupo = pd.DataFrame(grupo)

        # Criar um nome para a aba combinando os valores de "Nome Empto." e "Local.neg."
        nome_aba = f"{nome_empto[0]} 00{nome_empto[1]}"

        # Adicionar o DataFrame como uma nova aba no arquivo Excel
        with pd.ExcelWriter(caminho_arquivo, mode="a", engine="openpyxl") as writer:
            df_grupo.to_excel(writer, sheet_name=nome_aba, index=False)

    print("Tabelas separadas em abas com sucesso!")

    # Caminho do arquivo xlsx
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar o workbook
    workbook = openpyxl.load_workbook(file_path)

    # Iterar sobre as planilhas no workbook
    for sheet in workbook.sheetnames:
        # Obter a planilha atual
        current_sheet = workbook[sheet]
        # Ajustar tanto a coluna F quanto a coluna G, K, L e M
        for coluna_a_ajustar in ["F", "G", "K", "L", "M"]:
            # Iterar sobre as células na coluna a ser ajustada
            for row in range(2, current_sheet.max_row + 1):
                cell = current_sheet[coluna_a_ajustar + str(row)]
                # Aplicar formato contábil com números negativos entre parênteses
                cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

    # Salvar o arquivo
    workbook.save(file_path)

    # Carregar o arquivo Excel
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    workbook = load_workbook(caminho_arquivo)

    # Iterar sobre todas as abas
    for sheet_name in workbook.sheetnames:
        # Selecionar a aba
        sheet = workbook[sheet_name]

        # Inserir uma nova linha acima da primeira linha
        sheet.insert_rows(1)

    # Salvar as alterações
    workbook.save(caminho_arquivo)

    # Carregar o arquivo Excel
    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    wb = openpyxl.load_workbook(caminho_arquivo)

    # Iterar sobre todas as abas
    for sheet_name in wb.sheetnames:
        # Selecionar a aba
        planilha = wb[sheet_name]

        # Remover linhas de grade
        planilha.sheet_view.showGridLines = False

        # Congelar a quarta linha (linha 4, ou seja, A4)
        planilha.freeze_panes = "A3"

        # Definir larguras específicas para as colunas
        # Corrigir a estrutura do dicionário para definir corretamente as larguras das colunas D e E
        larguras = {"D": 20, "E": 20, "O": 15, "P": 20, "Q": 20}

        # Adicionar larguras padrão de 15 para as demais colunas, começando da coluna 2
        for coluna in range(2, planilha.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(coluna)
            if col_letter not in larguras:
                larguras[col_letter] = 13

        # Aplicar as larguras das colunas
        for col, width in larguras.items():
            planilha.column_dimensions[col].width = width

    # Salvar as alterações no arquivo
    wb.save(caminho_arquivo)

    # Carregar o arquivo Excel
    workbook = load_workbook(
        filename=r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    )

    # Iterar sobre todas as abas
    for sheet_name in workbook.sheetnames:
        # Selecionar a aba
        sheet = workbook[sheet_name]

        # Definir estilo para as células
        font = Font(bold=True, color="FF0000")  # Fonte em negrito e vermelha
        alignment = Alignment(horizontal="center")
        fill = PatternFill(
            start_color="FFffffff", end_color="FFffffff", fill_type="solid"
        )
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Mesclar células e escrever texto nelas com o estilo definido
        sheet.merge_cells("A1:Q1")
        cell_s4 = sheet["A1"]
        cell_s4.value = (
            "OBSERVAR LOCAL DA PRESTAÇÃO DE SERVIÇO (MATRIZ LOCALIZADA EM JOINVILLE)"
        )
        cell_s4.alignment = alignment
        cell_s4.font = font
        cell_s4.fill = fill
        cell_s4.border = border

    # Salvar as alterações no arquivo
    workbook.save(
        filename=r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    )

    # Carregar o arquivo Excel
    workbook = load_workbook(
        filename=r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    )

    # Iterar sobre todas as abas
    for sheet_name in workbook.sheetnames:
        # Selecionar a aba
        sheet = workbook[sheet_name]

        # Definir estilo para as células
        bold_font = Font(bold=True)
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Aplicar estilos à segunda linha
        for cell in sheet[2]:
            cell.font = bold_font
            cell.border = border_style

    # Salvar as alterações no arquivo
    workbook.save(
        filename=r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    )

    caminho_arquivo = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"
    workbook = load_workbook(caminho_arquivo)

    # Excluir a aba "Sheet1", se existir
    if "Sheet1" in workbook.sheetnames:
        del workbook["Sheet1"]

    # Iterar sobre todas as abas
    for sheet_name in workbook.sheetnames:
        # Selecionar a aba
        sheet = workbook[sheet_name]

        # Excluir a décima oitava coluna (R)
        sheet.delete_cols(18)

    # Salvar as alterações no arquivo
    workbook.save(caminho_arquivo)

    # Caminho para o arquivo Excel
    file_path = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Carregar a planilha
    workbook = openpyxl.load_workbook(file_path)

    # Obter os nomes das abas
    sheetnames = workbook.sheetnames

    # Ordenar os nomes das abas baseando-se nos dois últimos dígitos da direita para a esquerda
    sorted_sheetnames = sorted(sheetnames, key=lambda name: name[-2:])

    # Criar uma nova lista de abas ordenadas
    sorted_sheets = [workbook[name] for name in sorted_sheetnames]

    # Reordenar as abas no workbook
    workbook._sheets = sorted_sheets

    # Salvar o workbook ordenado
    workbook.save(file_path)

    print(f"Abas do arquivo {file_path} ordenadas com sucesso!")

    # Caminho do arquivo original
    caminho_original = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Caminho da pasta de destino
    pasta_destino = input_folder

    # Obtém apenas o dia da data atual
    dia_atual = data_atual.strftime("%d")

    # Nome do arquivo baseado no dia atual
    nome_arquivo = f"{dia_atual} Relatório de Conferência.xlsx"

    # Caminho completo do arquivo de destino
    caminho_destino = os.path.join(pasta_destino, nome_arquivo)

    # Copiando o arquivo
    shutil.copy(caminho_original, caminho_destino)

    # Caminhos dos arquivos a serem deletados
    caminho_matriz = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\\MATRIZ.xlsx"
    caminho_zfi017 = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência\ZFI017.xlsx"

    # Deletar os arquivos
    for caminho in [caminho_matriz, caminho_zfi017]:
        if os.path.exists(caminho):
            os.remove(caminho)
            print(f"Arquivo deletado com sucesso: {caminho}")
        else:
            print(f"O arquivo não foi encontrado: {caminho}")

    pasta_origem = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\Planilha de Conferência"
    pasta_destino = r"G:\Drives compartilhados\Fiscal_Arquivo_de_Notas\FISCAL_SERVIÇOS\CONFERENCIA Fiscal"
    organizar_arquivos(pasta_origem, pasta_destino)
    messagebox.showinfo("Finalizado", "Relatório gerado com sucesso!")


# Criando a janela principal
root = tk.Tk()
root.title("Execução de Rotinas")
root.geometry("250x300")  # Ajuste para comportar os widgets
root.configure(bg="#f2f2f2")
root.resizable(False, False)  # Impede o redimensionamento da janela

# Fonte personalizada
root.option_add("*Font", "Amiko 10")

# Título centralizado com quebra de linha
label_titulo = tk.Label(
    root,
    text="Relatório de Conferência ZFI017",
    font=("Amiko", 10, "bold"),
    bg="#f2f2f2",
    fg="#b23a48",
    wraplength=190,  # Ajuste a largura conforme necessário
)
label_titulo.pack(pady=(10, 0), padx=10)

# Frame principal para centralizar os elementos
frame = tk.Frame(root, bg="#f2f2f2")
frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

# Configuração de estilo para os widgets
input_style = {
    "font": ("Amiko", 10),
    "bg": "#ffffff",
    "bd": 0,
    "highlightthickness": 1,
    "highlightbackground": "#d1d1d1",
    "highlightcolor": "#b23a48",
    "relief": "flat",
}

label_style = {"bg": "#f2f2f2", "font": ("Amiko", 9)}
button_style = {
    "font": ("Amiko", 10),
    "bg": "#b23a48",
    "fg": "white",
    "bd": 0,
    "activebackground": "#8c2e39",
    "relief": "flat",
    "cursor": "hand2",
}

# Configurando os rótulos e entradas
tk.Label(frame, text="Usuário (SAP)", **label_style).grid(
    row=0, column=0, sticky="w", pady=(2, 2)
)
entry_usuario = tk.Entry(frame, **input_style)
entry_usuario.grid(row=0, column=1, sticky="ew", pady=(2, 8), padx=2)

tk.Label(frame, text="Senha (SAP)", **label_style).grid(
    row=1, column=0, sticky="w", pady=(2, 2)
)
entry_senha = tk.Entry(frame, **input_style, show="*")
entry_senha.grid(row=1, column=1, sticky="ew", pady=(2, 8), padx=2)

tk.Label(frame, text="Local - De", **label_style).grid(
    row=2, column=0, sticky="w", pady=(2, 2)
)
entry_loc1 = tk.Entry(frame, **input_style)
entry_loc1.grid(row=2, column=1, sticky="ew", pady=(2, 8), padx=2)

tk.Label(frame, text="Local - Até", **label_style).grid(
    row=3, column=0, sticky="w", pady=(2, 2)
)
entry_loc2 = tk.Entry(frame, **input_style)
entry_loc2.grid(row=3, column=1, sticky="ew", pady=(2, 8), padx=2)

tk.Label(frame, text="Data Inicio", **label_style).grid(
    row=4, column=0, sticky="w", pady=(2, 2)
)
entry_data1 = tk.Entry(frame, **input_style)
entry_data1.grid(row=4, column=1, sticky="ew", pady=(2, 8), padx=2)

tk.Label(frame, text="Data Fim", **label_style).grid(
    row=5, column=0, sticky="w", pady=(2, 2)
)
entry_data2 = tk.Entry(frame, **input_style)
entry_data2.grid(row=5, column=1, sticky="ew", pady=(2, 8), padx=2)

# Centralizar o botão
button_executar = tk.Button(
    frame, text="Gerar Relatório", command=executar_rotina, **button_style, width=20
)
button_executar.grid(row=6, column=0, columnspan=2, pady=(10, 20), sticky="ew")


# Configurar o comportamento de hover para os botões
def on_enter(e):
    e.widget["bg"] = "#8c2e39"


def on_leave(e):
    e.widget["bg"] = "#b23a48"


button_executar.bind("<Enter>", on_enter)
button_executar.bind("<Leave>", on_leave)


# Adicionando a funcionalidade de tecla Enter
def on_enter_pressed(event):
    print("Relatório gerado")


root.bind("<Return>", on_enter_pressed)

# Inicializa a interface
root.mainloop()
